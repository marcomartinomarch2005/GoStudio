"""
GoStudio — Unified Encode + Upload Pipeline
Single-file PySide6 app combining GoEncoder + GoUploader
4-Column UI: Media | Render+YouTube | Tracklist | Pipeline Queue+Log
"""

import sys
import os
import re
import json
import random
import threading
import datetime
import tempfile
import subprocess
import difflib
import time
import logging
from pathlib import Path
from zoneinfo import ZoneInfo

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QComboBox, QRadioButton, QButtonGroup,
    QLineEdit, QTextEdit, QFileDialog, QMessageBox, QFrame,
    QSizePolicy, QScrollArea, QSpacerItem, QGraphicsDropShadowEffect,
    QListWidget, QListWidgetItem, QGridLayout, QCheckBox,
    QPlainTextEdit, QProgressBar
)
from PySide6.QtCore import Qt, Signal, QObject, QEvent, QRectF
from PySide6.QtGui import (
    QFont, QPainter, QColor, QPen, QPalette, QPixmap, QPainterPath,
    QDrag, QImage
)

import imageio_ffmpeg
from moviepy import AudioFileClip

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS & CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

SCOPES = [
    "https://www.googleapis.com/auth/youtube.upload",
    "https://www.googleapis.com/auth/youtube",
    "https://www.googleapis.com/auth/youtube.force-ssl",
]
APP_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
TOKENS_DIR = APP_DIR / "tokens"
CLIENT_SECRET = APP_DIR / "client_secret.json"
LOGS_DIR = APP_DIR / "logs"
THUMBNAIL_CACHE_DIR = APP_DIR / "thumbnail_cache"
MAX_LOG_FILES = 7
WIT_TZ = ZoneInfo("Asia/Jayapura")  # UTC+9

YOUTUBE_CATEGORIES = {
    "Film & Animation": "1", "Autos & Vehicles": "2", "Music": "10",
    "Pets & Animals": "15", "Sports": "17", "Gaming": "20",
    "People & Blogs": "22", "Comedy": "23", "Entertainment": "24",
    "News & Politics": "25", "Howto & Style": "26", "Education": "27",
    "Science & Technology": "28", "Nonprofits & Activism": "29",
}

AUDIO_EXTENSIONS = {'.mp3', '.wav', '.m4a', '.aac', '.flac', '.ogg', '.wma'}
VIDEO_EXTENSIONS = {'.mp4', '.mkv', '.avi', '.mov', '.webm'}
IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png'}


# ═══════════════════════════════════════════════════════════════════════════════
# GPU ENCODER DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def deteksi_gpu_encoder():
    encoder_tersedia = []
    try:
        ffmpeg_exe = imageio_ffmpeg.get_ffmpeg_exe()
        hasil = subprocess.run(
            [ffmpeg_exe, "-hide_banner", "-encoders"],
            capture_output=True, text=True, encoding='utf-8', errors='ignore',
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        output = hasil.stdout
        if "h264_nvenc" in output:
            encoder_tersedia.append(("NVIDIA NVENC", "h264_nvenc"))
        if "h264_amf" in output:
            encoder_tersedia.append(("AMD AMF", "h264_amf"))
        if "h264_qsv" in output:
            encoder_tersedia.append(("Intel QSV", "h264_qsv"))
    except Exception:
        pass
    encoder_tersedia.append(("CPU (libx264)", "libx264"))
    return encoder_tersedia


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL IMPORT LOGIC (Sequence Variations + Visual SEO)
# ═══════════════════════════════════════════════════════════════════════════════

_COL_ALIASES = {
    'variation': ['variation', 'variasi', 'version', 'group', 'variant', 'variation name'],
    'track_order': ['track order', 'order', 'no', 'urutan', 'track', 'sequence'],
    'title': ['title', 'judul', 'nama', 'working title', 'song', 'lagu',
              'track title', 'track name', 'new track name', 'primary title', 'primary song title'],
    'session': ['session', 'sesi', 'section', 'session track', 'session function'],
    'filename': ['filename', 'file', 'nama file', 'file name', 'path'],
    'vocal': ['vocal', 'vokal', 'singer', 'voice'],
    'mood': ['mood', 'suasana', 'feel'],
    'note': ['strategic note', 'note', 'catatan', 'notes', 'strategic use', 'production note'],
}

_SEO_COL_ALIASES = {
    'variation': ['variation', 'variasi', 'version', 'variant', 'variation name'],
    'seo_title': ['title', 'judul', 'youtube title', 'video title', 'seo title'],
    'seo_description': ['description', 'deskripsi', 'desc', 'youtube description', 'description + hashtags'],
    'seo_tags': ['tags', 'tag', 'keywords', 'keyword', 'youtube tags'],
    'seo_hashtags': ['hashtags', 'hashtag', '#'],
}


def _detect_columns(header_row, aliases=None):
    if aliases is None:
        aliases = _COL_ALIASES
    mapping = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        val = str(cell).strip().lower()
        for key, alias_list in aliases.items():
            if val in alias_list:
                if key in mapping:
                    existing_val = str(header_row[mapping[key]]).strip().lower() if header_row[mapping[key]] else ""
                    if len(val) > len(existing_val):
                        mapping[key] = idx
                else:
                    mapping[key] = idx
                break
    return mapping


def _parse_excel_variations(filepath):
    """Parse Sequence Variations sheet."""
    if not HAS_OPENPYXL:
        return {'error': 'openpyxl tidak terinstall. Jalankan: pip install openpyxl'}
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return {'error': f'Gagal membaca Excel: {e}'}

    target_ws = None
    col_map = None

    # Priority 1: sheet with 'variation' + 'title' + 'track_order' (most specific = Sequence Variations)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        first_row = next(rows, None)
        if first_row is None:
            continue
        detected = _detect_columns(first_row)
        if 'variation' in detected and 'title' in detected and 'track_order' in detected:
            target_ws = ws
            col_map = detected
            break

    # Priority 2: any sheet with 'variation' + 'title' in header
    if target_ws is None:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            first_row = next(rows, None)
            if first_row is None:
                continue
            detected = _detect_columns(first_row)
            if 'variation' in detected and 'title' in detected:
                target_ws = ws
                col_map = detected
                break

    # Fallback: try first sheet with just 'title' column
    if target_ws is None:
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        first_row = next(rows, None)
        if first_row:
            detected = _detect_columns(first_row)
            if 'title' in detected:
                target_ws = ws
                col_map = detected

    if target_ws is None or col_map is None:
        wb.close()
        return {'error': 'Tidak ditemukan kolom "Variation" dan "Title" di Excel.'}

    variations = {}
    variation_names_ordered = []
    all_titles = set()
    var_num_map = {}  # {variation_name: variation_number_str} for SEO cross-ref

    # Detect the numeric "Variation" column separately (col_map['variation'] points to "Variation Name")
    rows_for_header = target_ws.iter_rows(values_only=True)
    header_row_seq = next(rows_for_header, None)
    var_num_col_idx = None
    if header_row_seq:
        for idx, cell in enumerate(header_row_seq):
            if cell is None:
                continue
            val = str(cell).strip().lower()
            if val in ('variation', 'variasi') and idx != col_map.get('variation'):
                var_num_col_idx = idx
                break
        # If col_map['variation'] actually IS the numeric column (no "Variation Name" found),
        # check if the data in first data row is numeric
        if var_num_col_idx is None and 'variation' in col_map:
            var_num_col_idx = None  # will try to extract number from var_name itself

    for row in target_ws.iter_rows(min_row=2, values_only=True):
        var_name = None
        if 'variation' in col_map:
            val = row[col_map['variation']] if col_map['variation'] < len(row) else None
            if val:
                var_name = str(val).strip()
        if not var_name:
            continue

        # Also grab variation number if separate column exists
        var_num_val = None
        if var_num_col_idx is not None and var_num_col_idx < len(row) and row[var_num_col_idx]:
            var_num_val = str(row[var_num_col_idx]).strip()

        title = None
        if 'title' in col_map:
            val = row[col_map['title']] if col_map['title'] < len(row) else None
            if val:
                title = str(val).strip()
        if not title:
            continue

        track_order = None
        if 'track_order' in col_map:
            val = row[col_map['track_order']] if col_map['track_order'] < len(row) else None
            if val is not None:
                try:
                    track_order = int(val)
                except (ValueError, TypeError):
                    track_order = None

        note = None
        if 'note' in col_map:
            val = row[col_map['note']] if col_map['note'] < len(row) else None
            if val:
                note = str(val).strip()

        track = {'title': title, 'track_order': track_order}

        if var_name not in variations:
            variations[var_name] = {'note': note or '', 'tracks': []}
            variation_names_ordered.append(var_name)
            if var_num_val:
                var_num_map[var_name] = var_num_val
        elif note and not variations[var_name]['note']:
            variations[var_name]['note'] = note

        variations[var_name]['tracks'].append(track)
        all_titles.add(title)

    wb.close()

    for var_name in variations:
        tracks = variations[var_name]['tracks']
        has_order = any(t['track_order'] is not None for t in tracks)
        if has_order:
            tracks.sort(key=lambda t: t['track_order'] if t['track_order'] is not None else 9999)

    if not variations:
        return {'error': 'Tidak ada data variasi yang valid di Excel.'}

    return {
        'variations': variations,
        'variation_names': variation_names_ordered,
        'all_titles': list(all_titles),
        'var_num_map': var_num_map,
        'error': None,
    }


def _parse_visual_seo(filepath):
    """Parse Visual SEO sheet for YouTube auto-fill data.
    Returns (dict, ordered_list). Dict keyed by variation name. List ordered by appearance."""
    if not HAS_OPENPYXL:
        return {}, []
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception:
        return {}, []

    seo_data = {}
    seo_ordered = []

    # Search ALL sheets for one that has columns matching SEO pattern
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        header = rows[0]
        if header is None:
            continue

        # Detect columns using SEO aliases
        col_map = _detect_columns(header, _SEO_COL_ALIASES)
        # Must have at least title + (description or tags) to qualify as SEO sheet
        if 'seo_title' not in col_map:
            continue
        if 'seo_description' not in col_map and 'seo_tags' not in col_map:
            continue

        # Found a valid SEO sheet — determine variation name column
        var_col = col_map.get('variation')

        for row in rows[1:]:
            if row is None or all(c is None for c in row):
                continue

            # Get variation name
            var_name = None
            if var_col is not None and var_col < len(row) and row[var_col]:
                var_name = str(row[var_col]).strip()
            if not var_name:
                continue

            # Get title
            seo_title = ""
            if 'seo_title' in col_map and col_map['seo_title'] < len(row) and row[col_map['seo_title']]:
                seo_title = str(row[col_map['seo_title']]).strip()

            # Get description
            seo_desc = ""
            if 'seo_description' in col_map and col_map['seo_description'] < len(row) and row[col_map['seo_description']]:
                seo_desc = str(row[col_map['seo_description']]).strip()

            # Get tags
            seo_tags = ""
            if 'seo_tags' in col_map and col_map['seo_tags'] < len(row) and row[col_map['seo_tags']]:
                seo_tags = str(row[col_map['seo_tags']]).strip()

            # Get hashtags (separate column if exists)
            seo_hashtags = ""
            if 'seo_hashtags' in col_map and col_map['seo_hashtags'] < len(row) and row[col_map['seo_hashtags']]:
                seo_hashtags = str(row[col_map['seo_hashtags']]).strip()

            # Combine description + hashtags
            full_desc = seo_desc
            if seo_hashtags and seo_hashtags not in seo_desc:
                full_desc = f"{seo_desc}\n\n{seo_hashtags}" if seo_desc else seo_hashtags

            entry = {'title': seo_title, 'description': full_desc, 'tags': seo_tags}
            seo_data[var_name] = entry
            seo_ordered.append(entry)

        # Found and parsed — stop looking at other sheets
        if seo_ordered:
            break

    wb.close()
    return seo_data, seo_ordered


def _scan_audio_folder(folder_path):
    files = []
    if not os.path.isdir(folder_path):
        return files
    for f in os.listdir(folder_path):
        ext = os.path.splitext(f)[1].lower()
        if ext in AUDIO_EXTENSIONS:
            files.append(os.path.join(folder_path, f))
    return sorted(files)


def _strip_number_prefix(name):
    return re.sub(r'^\d+[\.\-_\s]+\s*', '', name).strip()


def _match_title_to_files(title, audio_files, threshold=0.6):
    title_clean = title.lower().strip()
    best_match = None
    best_ratio = 0
    for fpath in audio_files:
        fname = os.path.splitext(os.path.basename(fpath))[0]
        fname_clean = _strip_number_prefix(fname).lower().strip()
        if fname_clean == title_clean:
            return fpath
        if title_clean in fname_clean or fname_clean in title_clean:
            ratio = 0.95
        else:
            ratio = difflib.SequenceMatcher(None, title_clean, fname_clean).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = fpath
    if best_ratio >= threshold:
        return best_match
    return None


def _match_all_titles(variations_data, audio_files):
    all_titles = variations_data['all_titles']
    matches = {}
    for title in all_titles:
        match = _match_title_to_files(title, audio_files)
        matches[title] = match
    return matches


# ═══════════════════════════════════════════════════════════════════════════════
# YOUTUBE AUTH MANAGER
# ═══════════════════════════════════════════════════════════════════════════════

class YouTubeAuth:
    @staticmethod
    def get_channels():
        if not TOKENS_DIR.exists():
            TOKENS_DIR.mkdir(parents=True, exist_ok=True)
        return [f.stem for f in TOKENS_DIR.glob("*.json") if f.name != "client_secret.json"]

    @staticmethod
    def authenticate(channel_name):
        if not CLIENT_SECRET.exists():
            return None, "client_secret.json tidak ditemukan!"
        token_path = TOKENS_DIR / f"{channel_name}.json"
        creds = None
        if token_path.exists():
            creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())
                except Exception:
                    creds = None
            if not creds:
                flow = InstalledAppFlow.from_client_secrets_file(str(CLIENT_SECRET), SCOPES)
                creds = flow.run_local_server(port=0)
            with open(token_path, 'w') as f:
                f.write(creds.to_json())
        youtube = build("youtube", "v3", credentials=creds)
        return youtube, None

    @staticmethod
    def get_service(channel_name):
        token_path = TOKENS_DIR / f"{channel_name}.json"
        if not token_path.exists():
            return None, "Token tidak ditemukan. Login ulang."
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())
                    with open(token_path, 'w') as f:
                        f.write(creds.to_json())
                except Exception as e:
                    return None, f"Gagal refresh token: {e}"
            else:
                return None, "Token expired. Login ulang."
        youtube = build("youtube", "v3", credentials=creds)
        return youtube, None

    @staticmethod
    def get_channel_info(youtube):
        try:
            resp = youtube.channels().list(part="snippet", mine=True).execute()
            if resp.get("items"):
                item = resp["items"][0]
                return item["snippet"]["title"], item["id"], item["snippet"].get("thumbnails", {}).get("default", {}).get("url", "")
        except Exception:
            pass
        return None, None, ""

    @staticmethod
    def get_playlists(youtube):
        playlists = []
        try:
            next_page = None
            while True:
                resp = youtube.playlists().list(part="snippet", mine=True, maxResults=50, pageToken=next_page).execute()
                for item in resp.get("items", []):
                    playlists.append({"id": item["id"], "title": item["snippet"]["title"]})
                next_page = resp.get("nextPageToken")
                if not next_page:
                    break
        except Exception:
            pass
        return playlists

    @staticmethod
    def remove_channel(channel_name):
        token_path = TOKENS_DIR / f"{channel_name}.json"
        if token_path.exists():
            token_path.unlink()


# ═══════════════════════════════════════════════════════════════════════════════
# UPLOAD FUNCTION (Resumable 10MB chunks)
# ═══════════════════════════════════════════════════════════════════════════════

def upload_video_to_youtube(youtube, task, signals, task_index):
    """Upload a single video to YouTube with resumable upload. Returns video_id or None."""
    try:
        video_path = task["video_path"]
        title = task["yt_title"]
        description = task.get("yt_description", "")
        tags = task.get("yt_tags", [])
        category_id = task.get("yt_category_id", "10")
        privacy = task.get("yt_privacy", "private")
        scheduled_time = task.get("yt_scheduled_time")
        thumbnail_path = task.get("yt_thumbnail_path")
        playlist_id = task.get("yt_playlist_id")

        body = {
            "snippet": {"title": title, "description": description, "tags": tags, "categoryId": category_id},
            "status": {"privacyStatus": privacy, "selfDeclaredMadeForKids": False},
        }
        if scheduled_time and privacy == "private":
            body["status"]["publishAt"] = scheduled_time

        signals.log.emit(f"[Upload] Mulai: {title}")

        media = MediaFileUpload(video_path, mimetype="video/mp4", resumable=True, chunksize=10 * 1024 * 1024)
        request = youtube.videos().insert(part="snippet,status", body=body, media_body=media)

        response = None
        while response is None:
            status, response = request.next_chunk()
            if status:
                pct = status.progress() * 100
                signals.upload_progress.emit(task_index, pct)

        video_id = response.get("id", "")
        signals.log.emit(f"[Upload] Selesai! ID: {video_id} | https://youtu.be/{video_id}")

        # Set thumbnail
        if thumbnail_path and os.path.exists(thumbnail_path):
            try:
                youtube.thumbnails().set(videoId=video_id, media_body=MediaFileUpload(thumbnail_path)).execute()
                signals.log.emit(f"[Upload] Thumbnail set.")
            except Exception as e:
                signals.log.emit(f"[Upload] Thumbnail gagal: {e}")

        # Add to playlist
        if playlist_id:
            try:
                youtube.playlistItems().insert(
                    part="snippet",
                    body={"snippet": {"playlistId": playlist_id, "resourceId": {"kind": "youtube#video", "videoId": video_id}}}
                ).execute()
                signals.log.emit(f"[Upload] Ditambahkan ke playlist.")
            except Exception as e:
                signals.log.emit(f"[Upload] Playlist gagal: {e}")

        return video_id
    except Exception as e:
        signals.log.emit(f"[Upload] ERROR: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# FORCE LIGHT PALETTE (Teal/Emerald theme)
# ═══════════════════════════════════════════════════════════════════════════════

def force_light_palette(app):
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor("#F0FDFA"))
    palette.setColor(QPalette.WindowText, QColor("#1a1a2e"))
    palette.setColor(QPalette.Base, QColor("#FFFFFF"))
    palette.setColor(QPalette.AlternateBase, QColor("#F7FAFA"))
    palette.setColor(QPalette.Text, QColor("#1a1a2e"))
    palette.setColor(QPalette.Button, QColor("#F1F5F4"))
    palette.setColor(QPalette.ButtonText, QColor("#1a1a2e"))
    palette.setColor(QPalette.Highlight, QColor("#0D9488"))
    palette.setColor(QPalette.HighlightedText, QColor("#FFFFFF"))
    palette.setColor(QPalette.ToolTipBase, QColor("#FFFFFF"))
    palette.setColor(QPalette.ToolTipText, QColor("#1a1a2e"))
    palette.setColor(QPalette.PlaceholderText, QColor("#9CA3AF"))
    app.setPalette(palette)


# ═══════════════════════════════════════════════════════════════════════════════
# GLOBAL STYLESHEET
# ═══════════════════════════════════════════════════════════════════════════════

STYLESHEET = """
* { font-family: 'Segoe UI', sans-serif; font-size: 13px; }
QMainWindow, QWidget { background-color: #F0FDFA; color: #1a1a2e; }
QLabel { color: #1a1a2e; background: transparent; border: none; }
QRadioButton { color: #1a1a2e; font-size: 13px; spacing: 10px; background: transparent; border: none; }
QRadioButton::indicator { width: 20px; height: 20px; border-radius: 10px; border: 2px solid #D1D5DB; background-color: #FFFFFF; }
QRadioButton::indicator:checked { border: 3px solid #0D9488; background-color: #0D9488; }
QRadioButton::indicator:hover { border-color: #5EEAD4; }
QComboBox { background-color: #FAFBFC; border: 1px solid #E5E7EB; border-radius: 10px; padding: 9px 30px 9px 12px; font-size: 13px; color: #1a1a2e; }
QComboBox:hover { border-color: #0D9488; }
QComboBox::drop-down { border: none; width: 28px; background: transparent; }
QComboBox::down-arrow { image: none; border: none; width: 0; height: 0; }
QComboBox QAbstractItemView { background-color: #FFFFFF; border: none; outline: none; color: #1a1a2e; }
QComboBox QAbstractItemView::item { padding: 8px 12px; border: none; border-radius: 6px; min-height: 28px; }
QComboBox QAbstractItemView::item:selected { background-color: #F0FDFA; color: #1a1a2e; }
QLineEdit { background-color: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 8px; padding: 7px 10px; font-size: 13px; color: #1a1a2e; }
QLineEdit:focus { border-color: #0D9488; }
QTextEdit, QPlainTextEdit { background-color: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 8px; padding: 7px 10px; font-size: 13px; color: #1a1a2e; }
QTextEdit:focus, QPlainTextEdit:focus { border-color: #0D9488; }
QScrollBar:vertical { background: transparent; width: 7px; margin: 0; }
QScrollBar::handle:vertical { background: #D1D5DB; border-radius: 3px; min-height: 30px; }
QScrollBar::handle:vertical:hover { background: #9CA3AF; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: transparent; }
QScrollBar:horizontal { background: transparent; height: 7px; margin: 0; }
QScrollBar::handle:horizontal { background: #D1D5DB; border-radius: 3px; min-width: 30px; }
QScrollBar::handle:horizontal:hover { background: #9CA3AF; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal { background: transparent; }
QCheckBox { color: #1a1a2e; font-size: 13px; spacing: 10px; background: transparent; border: none; }
QCheckBox::indicator { width: 20px; height: 20px; border-radius: 6px; border: 2px solid #D1D5DB; background-color: #FFFFFF; }
QCheckBox::indicator:checked { border: 3px solid #0D9488; background-color: #0D9488; }
QCheckBox::indicator:hover { border-color: #5EEAD4; }
QProgressBar { background-color: #E5E7EB; border: none; border-radius: 3px; height: 6px; }
QProgressBar::chunk { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #0D9488,stop:1 #10B981); border-radius: 3px; }
"""


# ═══════════════════════════════════════════════════════════════════════════════
# WORKER SIGNALS
# ═══════════════════════════════════════════════════════════════════════════════

class PipelineSignals(QObject):
    log = Signal(str)
    encode_progress = Signal(int, float)   # task_index, pct
    upload_progress = Signal(int, float)   # task_index, pct
    encode_done = Signal(int, bool)        # task_index, success
    upload_done = Signal(int, bool)        # task_index, success
    status_changed = Signal(int, str)      # task_index, new_status


# ═══════════════════════════════════════════════════════════════════════════════
# CUSTOM FRAMELESS COMBOBOX POPUP
# ═══════════════════════════════════════════════════════════════════════════════

class ComboPopupFrame(QFrame):
    def __init__(self, combo_box):
        super().__init__(None)
        self._combo = combo_box
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Tool | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setAttribute(Qt.WA_DeleteOnClose)
        self._setup_ui()
        QApplication.instance().installEventFilter(self)

    def _setup_ui(self):
        container = QFrame(self)
        container.setObjectName("popupContainer")
        container.setStyleSheet("QFrame#popupContainer { background: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 10px; }")
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(24)
        shadow.setColor(QColor(0, 0, 0, 40))
        shadow.setOffset(0, 4)
        container.setGraphicsEffect(shadow)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(container)
        layout = QVBoxLayout(container)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(2)
        self._list = QListWidget()
        self._list.setFrameShape(QFrame.NoFrame)
        self._list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._list.setStyleSheet(
            "QListWidget { background: transparent; border: none; outline: none; padding: 2px; }"
            "QListWidget::item { padding: 8px 12px; border: none; border-radius: 6px; min-height: 22px; color: #1a1a2e; font-size: 13px; }"
            "QListWidget::item:hover { background: #F0FDFA; }"
            "QListWidget::item:selected { background: #CCFBF1; color: #1a1a2e; }"
        )
        for i in range(self._combo.count()):
            item = QListWidgetItem(self._combo.itemText(i))
            item.setData(Qt.UserRole, i)
            self._list.addItem(item)
            if i == self._combo.currentIndex():
                self._list.setCurrentItem(item)
        self._list.itemClicked.connect(self._on_item_clicked)
        layout.addWidget(self._list)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.MouseButtonPress:
            if not self.geometry().contains(event.globalPosition().toPoint()):
                self.close()
        return False

    def _on_item_clicked(self, item):
        self._combo.setCurrentIndex(item.data(Qt.UserRole))
        self.close()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.close()
        elif event.key() in (Qt.Key_Return, Qt.Key_Enter):
            current = self._list.currentItem()
            if current:
                self._on_item_clicked(current)
        elif event.key() == Qt.Key_Up:
            r = self._list.currentRow()
            if r > 0:
                self._list.setCurrentRow(r - 1)
        elif event.key() == Qt.Key_Down:
            r = self._list.currentRow()
            if r < self._list.count() - 1:
                self._list.setCurrentRow(r + 1)
        else:
            super().keyPressEvent(event)

    def closeEvent(self, event):
        QApplication.instance().removeEventFilter(self)
        super().closeEvent(event)

    def show_at(self):
        self.adjustSize()
        bottom_left = self._combo.mapToGlobal(self._combo.rect().bottomLeft())
        h = min(self._combo.count() * 38 + 16, 280)
        self.setGeometry(bottom_left.x(), bottom_left.y(), self._combo.width(), h)
        self.show()
        self._list.setFocus()
        self.raise_()


class ModernComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._popup = None

    def showPopup(self):
        if self._popup and self._popup.isVisible():
            self._popup.close()
            return
        if self.count() == 0:
            return
        self._popup = ComboPopupFrame(self)
        self._popup.destroyed.connect(self._on_popup_destroyed)
        self._popup.show_at()

    def _on_popup_destroyed(self):
        self._popup = None

    def hidePopup(self):
        if self._popup:
            self._popup.close()


# ═══════════════════════════════════════════════════════════════════════════════
# CIRCULAR PROGRESS WIDGET
# ═══════════════════════════════════════════════════════════════════════════════

class CircularProgress(QWidget):
    def __init__(self, size=44, parent=None):
        super().__init__(parent)
        self.setFixedSize(size, size)
        self._value = 0
        self._finished = False
        self._color = QColor("#0D9488")

    def set_value(self, val):
        self._value = val
        self._finished = False
        self.update()

    def set_finished(self):
        self._value = 100
        self._finished = True
        self.update()

    def set_color(self, color):
        self._color = QColor(color)
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        size = min(self.width(), self.height())
        rect_size = size - 10
        x = (self.width() - rect_size) / 2
        y = (self.height() - rect_size) / 2
        pen = QPen(QColor("#E5E7EB"), 4)
        pen.setCapStyle(Qt.RoundCap)
        painter.setPen(pen)
        painter.drawEllipse(int(x), int(y), rect_size, rect_size)
        color = QColor("#10B981") if self._finished else self._color
        pen = QPen(color, 4)
        pen.setCapStyle(Qt.RoundCap)
        painter.setPen(pen)
        span = int(-self._value * 360 / 100 * 16)
        painter.drawArc(int(x), int(y), rect_size, rect_size, 90 * 16, span)
        painter.setPen(color)
        font = QFont("Segoe UI", 8, QFont.Bold)
        painter.setFont(font)
        text = "\u2713" if self._finished else f"{int(self._value)}%"
        painter.drawText(self.rect(), Qt.AlignCenter, text)
        painter.end()


# ═══════════════════════════════════════════════════════════════════════════════
# AUDIO ITEM WIDGET + DRAG-DROP LIST
# ═══════════════════════════════════════════════════════════════════════════════

class AudioItemWidget(QWidget):
    def __init__(self, index, filename, duration_str, parent=None):
        super().__init__(parent)
        self.index = index
        self.setStyleSheet("background: transparent; border: none;")
        layout = QHBoxLayout(self)
        layout.setContentsMargins(8, 6, 8, 6)
        layout.setSpacing(8)

        self.checkbox = QCheckBox()
        self.checkbox.setStyleSheet(
            "QCheckBox::indicator { width: 16px; height: 16px; border-radius: 4px; border: 2px solid #D1D5DB; background-color: #FFFFFF; }"
            "QCheckBox::indicator:checked { border: 2px solid #0D9488; background-color: #0D9488; }"
        )
        layout.addWidget(self.checkbox)

        self.num_label = QLabel(str(index + 1))
        self.num_label.setFixedSize(24, 24)
        self.num_label.setAlignment(Qt.AlignCenter)
        self.num_label.setStyleSheet("background-color: #0D9488; color: white; border-radius: 6px; font-size: 10px; font-weight: bold;")
        layout.addWidget(self.num_label)

        info_layout = QVBoxLayout()
        info_layout.setSpacing(1)
        name_label = QLabel(filename)
        name_label.setStyleSheet("font-size: 11px; font-weight: 500; color: #1a1a2e; background: transparent;")
        dur_label = QLabel(duration_str)
        dur_label.setStyleSheet("font-size: 10px; color: #9CA3AF; background: transparent;")
        info_layout.addWidget(name_label)
        info_layout.addWidget(dur_label)
        layout.addLayout(info_layout, 1)

        drag_label = QLabel("\u2630")
        drag_label.setStyleSheet("font-size: 14px; color: #D1D5DB; background: transparent;")
        layout.addWidget(drag_label)


class DragDropAudioList(QWidget):
    order_changed = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._parent_window = parent
        self.setStyleSheet("background: transparent;")
        self._layout = QVBoxLayout(self)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(0)
        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self._container = QWidget()
        self._container.setStyleSheet("background: transparent;")
        self._list_layout = QVBoxLayout(self._container)
        self._list_layout.setContentsMargins(0, 0, 0, 0)
        self._list_layout.setSpacing(3)
        self._list_layout.addStretch()
        self._scroll.setWidget(self._container)
        self._layout.addWidget(self._scroll)
        self._items = []

    def clear(self):
        while self._list_layout.count() > 1:
            item = self._list_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._items.clear()

    def count(self):
        return len(self._items)

    def add_item(self, widget, index):
        frame = DraggableAudioFrame(self, index)
        frame_layout = QVBoxLayout(frame)
        frame_layout.setContentsMargins(0, 0, 0, 0)
        frame_layout.addWidget(widget)
        frame._audio_widget = widget
        self._list_layout.insertWidget(len(self._items), frame)
        self._items.append(frame)

    def get_widget(self, index):
        if 0 <= index < len(self._items):
            return self._items[index]._audio_widget
        return None

    def move_item(self, from_idx, to_idx):
        if from_idx == to_idx or from_idx < 0 or to_idx < 0:
            return
        if from_idx >= len(self._items) or to_idx >= len(self._items):
            return
        pw = self._parent_window
        if pw and hasattr(pw, 'daftar_audio'):
            audio = pw.daftar_audio.pop(from_idx)
            dur = pw.audio_durations.pop(from_idx)
            pw.daftar_audio.insert(to_idx, audio)
            pw.audio_durations.insert(to_idx, dur)
            pw._refresh_audio_list()
            self.order_changed.emit()


class DraggableAudioFrame(QFrame):
    STYLE_NORMAL = "QFrame { background-color: #FAFBFC; border: 1px solid #F0F0F0; border-radius: 8px; }" \
                   "QFrame:hover { background-color: #F0FDFA; border-color: #0D9488; }"
    STYLE_DRAGGING = "QFrame { background-color: #FEF3C7; border: 2px dashed #F59E0B; border-radius: 8px; }"

    def __init__(self, list_widget, index, parent=None):
        super().__init__(parent)
        self._list_widget = list_widget
        self._index = index
        self.setStyleSheet(self.STYLE_NORMAL)
        self.setAcceptDrops(True)
        self._drag_start = None

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_start = event.position().toPoint()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_start is None:
            return super().mouseMoveEvent(event)
        if (event.position().toPoint() - self._drag_start).manhattanLength() < 20:
            return super().mouseMoveEvent(event)
        from PySide6.QtCore import QMimeData, QPoint
        drag = QDrag(self)
        mime = QMimeData()
        mime.setText(str(self._index))
        drag.setMimeData(mime)
        pixmap = QPixmap(self.size())
        pixmap.fill(QColor(0, 0, 0, 0))
        self.render(pixmap)
        scaled = pixmap.scaled(int(self.width() * 0.95), int(self.height() * 0.95), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        drag.setPixmap(scaled)
        drag.setHotSpot(QPoint(scaled.width() // 2, scaled.height() // 2))
        self.setStyleSheet(self.STYLE_DRAGGING)
        drag.exec(Qt.MoveAction)
        self.setStyleSheet(self.STYLE_NORMAL)
        self._drag_start = None

    def dragEnterEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()

    def dragMoveEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()

    def dragLeaveEvent(self, event):
        self.setStyleSheet(self.STYLE_NORMAL)

    def dropEvent(self, event):
        from_idx = int(event.mimeData().text())
        to_idx = self._index
        self.setStyleSheet(self.STYLE_NORMAL)
        if from_idx != to_idx:
            self._list_widget.move_item(from_idx, to_idx)
        event.acceptProposedAction()


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def make_card():
    frame = QFrame()
    frame.setObjectName("card")
    frame.setStyleSheet("QFrame#card { background-color: #FFFFFF; border-radius: 14px; border: 1px solid #E0F2F1; }")
    layout = QVBoxLayout(frame)
    layout.setContentsMargins(14, 12, 14, 14)
    layout.setSpacing(8)
    return frame, layout


def make_card_title(text):
    lbl = QLabel(text.upper())
    lbl.setStyleSheet("font-size: 10px; font-weight: bold; color: #6B7280; letter-spacing: 0.6px; padding: 0; margin: 0; border: none;")
    lbl.setContentsMargins(0, 0, 0, 0)
    return lbl


def make_btn(text, bg="#F1F5F4", fg="#374151", bold=True):
    btn = QPushButton(text)
    btn.setCursor(Qt.PointingHandCursor)
    btn.setStyleSheet(
        f"QPushButton {{ background-color: {bg}; color: {fg}; border: none; border-radius: 8px;"
        f"padding: 7px 12px; font-size: 11px; font-weight: {'bold' if bold else 'normal'}; }}"
        f"QPushButton:hover {{ background-color: {_darken(bg)}; }}"
    )
    return btn


def _darken(hex_color):
    darken_map = {
        "#F1F5F4": "#E2E8E7", "#0D9488": "#0B7C72", "#F97316": "#EA6C0B",
        "#FFFFFF": "#F3F4F6", "#ECFDF5": "#D1FAE5", "#14B8A6": "#0D9488",
        "#F3F4F6": "#E5E7EB",
    }
    return darken_map.get(hex_color, "#E5E7EB")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN WINDOW — GoStudio
# ═══════════════════════════════════════════════════════════════════════════════

class GoStudioWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("GoStudio — Unified Pipeline")
        self.setMinimumSize(1440, 780)
        self.resize(1520, 840)

        # ── State ──
        self.jalur_video = ""
        self.daftar_audio = []
        self.audio_durations = []
        self.daftar_encoder = deteksi_gpu_encoder()

        # Pipeline queue: list of task dicts
        self.pipeline_queue = []
        self.pipeline_widgets = []  # UI widgets per task
        self._encode_slot = None   # index of currently encoding task
        self._upload_slot = None   # index of currently uploading task
        self._current_proc = None
        self._cancel_requested = set()

        # YouTube
        self.channels = {}
        self.channel_info = {}
        self.playlists = []

        # Excel
        self._excel_data = None
        self._excel_seo_data = {}
        self._excel_seo_ordered = []
        self._excel_folder = ""
        self._excel_matches = {}
        self._excel_filepath = ""
        self._audio_mode = 'manual'

        # Signals
        self.signals = PipelineSignals()
        self.signals.log.connect(self._on_log)
        self.signals.encode_progress.connect(self._on_encode_progress)
        self.signals.upload_progress.connect(self._on_upload_progress)
        self.signals.encode_done.connect(self._on_encode_done)
        self.signals.upload_done.connect(self._on_upload_done)
        self.signals.status_changed.connect(self._on_status_changed)

        self._build_ui()
        self._load_saved_channels()
        self._init_logging()

    # ─── Logging ─────────────────────────────────────────────────────────────
    def _init_logging(self):
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        log_files = sorted(LOGS_DIR.glob("GoStudio_*.log"), key=lambda f: f.stat().st_mtime, reverse=True)
        for f in log_files[MAX_LOG_FILES:]:
            try:
                f.unlink()
            except OSError:
                pass
        ts = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        self._log_file = open(LOGS_DIR / f'GoStudio_{ts}.log', 'w', encoding='utf-8')
        self._log_file.write(f"--- GoStudio Log: {datetime.datetime.now():%Y-%m-%d %H:%M:%S} ---\n")

    def closeEvent(self, event):
        if hasattr(self, '_log_file') and self._log_file:
            try:
                self._log_file.write(f"--- Sesi ditutup: {datetime.datetime.now():%Y-%m-%d %H:%M:%S} ---\n")
                self._log_file.close()
            except Exception:
                pass
        super().closeEvent(event)

    # ─── Load Channels ───────────────────────────────────────────────────────
    def _load_saved_channels(self):
        for name in YouTubeAuth.get_channels():
            self.combo_channel.addItem(name)
        if self.combo_channel.count() > 0:
            self._on_channel_changed(0)

    # ─── BUILD UI ────────────────────────────────────────────────────────────
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(16, 12, 16, 12)
        root.setSpacing(10)

        # ── Header ──
        header = QHBoxLayout()
        lbl_title = QLabel("Go<span style='color:#0D9488;'>Studio</span>")
        lbl_title.setTextFormat(Qt.RichText)
        lbl_title.setStyleSheet("font-size: 20px; font-weight: bold; color: #1a1a2e;")
        header.addWidget(lbl_title)
        header.addStretch()

        gpu_name = self.daftar_encoder[0][0] if self.daftar_encoder[0][1] != "libx264" else "CPU Only"
        badge = QLabel(f"\u26A1 {gpu_name}")
        badge.setStyleSheet("background-color: #ECFDF5; color: #059669; padding: 5px 12px; border-radius: 12px; font-size: 11px; font-weight: bold;")
        header.addWidget(badge)

        self.btn_reset = QPushButton("Reset")
        self.btn_reset.setStyleSheet(
            "QPushButton { background: #FFFFFF; color: #DC2626; border: 1px solid #FECACA; border-radius: 14px; padding: 5px 14px; font-size: 11px; font-weight: bold; }"
            "QPushButton:hover { background: #FEF2F2; }"
        )
        self.btn_reset.clicked.connect(self._reset_all)
        header.addWidget(self.btn_reset)
        root.addLayout(header)

        # ── 4-Column Layout ──
        columns = QHBoxLayout()
        columns.setSpacing(12)

        # Column 1: Media
        col1 = QVBoxLayout()
        col1.setSpacing(10)
        self._build_col1_media(col1)
        columns.addLayout(col1, 22)

        # Column 2: Render Options + YouTube Detail
        col2 = QVBoxLayout()
        col2.setSpacing(10)
        self._build_col2_render_youtube(col2)
        columns.addLayout(col2, 33)

        # Column 3: Tracklist/Timestamp
        col3 = QVBoxLayout()
        col3.setSpacing(10)
        self._build_col3_tracklist(col3)
        columns.addLayout(col3, 20)

        # Column 4: Pipeline Queue + Log
        col4 = QVBoxLayout()
        col4.setSpacing(10)
        self._build_col4_pipeline(col4)
        columns.addLayout(col4, 25)

        root.addLayout(columns, 1)


    # ═══ COLUMN 1: MEDIA ═════════════════════════════════════════════════════
    def _build_col1_media(self, col):
        # Video card
        c_video, l_video = make_card()
        l_video.addWidget(make_card_title("\U0001F3AC Video Utama"))
        self.btn_video = make_btn("Pilih Video", "#0D9488", "white")
        self.btn_video.clicked.connect(self._pilih_video)
        self.lbl_video = QLabel("Belum dipilih")
        self.lbl_video.setStyleSheet("font-size: 11px; color: #9CA3AF;")
        row_v = QHBoxLayout()
        row_v.addWidget(self.btn_video)
        row_v.addWidget(self.lbl_video, 1)
        l_video.addLayout(row_v)

        self.video_preview = QLabel("Drop atau pilih video")
        self.video_preview.setFixedHeight(100)
        self.video_preview.setAlignment(Qt.AlignCenter)
        self.video_preview.setStyleSheet("QLabel { background-color: #F7FAFA; color: #9CA3AF; border: 2px dashed #D1D5DB; border-radius: 10px; font-size: 11px; }")
        self.video_preview.setAcceptDrops(True)
        l_video.addWidget(self.video_preview)
        col.addWidget(c_video)

        # Audio card
        c_audio, l_audio = make_card()
        l_audio.addWidget(make_card_title("\U0001F3B5 Daftar Audio"))

        # Mode toggle
        mode_toggle_frame = QFrame()
        mode_toggle_frame.setStyleSheet("QFrame { background-color: #F3F4F6; border-radius: 8px; padding: 2px; }")
        mode_toggle_inner = QHBoxLayout(mode_toggle_frame)
        mode_toggle_inner.setContentsMargins(2, 2, 2, 2)
        mode_toggle_inner.setSpacing(2)
        self.btn_mode_manual = QPushButton("Manual")
        self.btn_mode_excel = QPushButton("\U0001F4CA Import Excel")
        for btn in [self.btn_mode_manual, self.btn_mode_excel]:
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setStyleSheet(
                "QPushButton { background: transparent; border: none; border-radius: 6px; padding: 5px 10px; font-size: 10px; font-weight: 600; color: #6B7280; }"
                "QPushButton:checked { background: #FFFFFF; color: #0D9488; }"
            )
        self.btn_mode_manual.setChecked(True)
        self.btn_mode_manual.clicked.connect(lambda: self._set_audio_mode('manual'))
        self.btn_mode_excel.clicked.connect(lambda: self._set_audio_mode('excel'))
        mode_toggle_inner.addWidget(self.btn_mode_manual)
        mode_toggle_inner.addWidget(self.btn_mode_excel)
        l_audio.addWidget(mode_toggle_frame)

        # Manual panel
        self.manual_audio_panel = QWidget()
        manual_layout = QVBoxLayout(self.manual_audio_panel)
        manual_layout.setContentsMargins(0, 6, 0, 0)
        manual_layout.setSpacing(6)
        toolbar = QHBoxLayout()
        toolbar.setSpacing(4)
        btn_add = make_btn("+ Tambah", "#0D9488", "white")
        btn_add.clicked.connect(self._tambah_audio)
        btn_del = make_btn("Hapus")
        btn_del.clicked.connect(self._hapus_audio)
        btn_shuffle = make_btn("\U0001F500 Acak")
        btn_shuffle.clicked.connect(self._acak_audio)
        for b in [btn_add, btn_del, btn_shuffle]:
            toolbar.addWidget(b)
        toolbar.addStretch()
        manual_layout.addLayout(toolbar)
        self.audio_list_widget = DragDropAudioList(self)
        self.audio_list_widget.order_changed.connect(self._on_audio_reordered)
        manual_layout.addWidget(self.audio_list_widget, 1)
        l_audio.addWidget(self.manual_audio_panel, 1)

        # Excel panel
        self.excel_audio_panel = QWidget()
        excel_layout = QVBoxLayout(self.excel_audio_panel)
        excel_layout.setContentsMargins(0, 6, 0, 0)
        excel_layout.setSpacing(8)
        self.excel_import_zone = QPushButton("\U0001F4CA  Klik untuk Import Excel (.xlsx)")
        self.excel_import_zone.setCursor(Qt.PointingHandCursor)
        self.excel_import_zone.setStyleSheet(
            "QPushButton { background: #F0FDFA; border: 2px dashed #0D9488; border-radius: 10px; padding: 14px; font-size: 11px; font-weight: 600; color: #6B7280; }"
            "QPushButton:hover { border-color: #0B7C72; background: #CCFBF1; }"
        )
        self.excel_import_zone.clicked.connect(self._import_excel)
        excel_layout.addWidget(self.excel_import_zone)

        self.excel_variation_widget = QWidget()
        excel_var_layout = QVBoxLayout(self.excel_variation_widget)
        excel_var_layout.setContentsMargins(0, 0, 0, 0)
        excel_var_layout.setSpacing(5)
        excel_var_layout.addWidget(QLabel("Variasi:"))
        self.combo_variation = ModernComboBox()
        self.combo_variation.currentIndexChanged.connect(self._on_variation_changed)
        excel_var_layout.addWidget(self.combo_variation)
        self.lbl_variation_note = QLabel("")
        self.lbl_variation_note.setWordWrap(True)
        self.lbl_variation_note.setStyleSheet("font-size: 10px; color: #6B7280; font-style: italic;")
        excel_var_layout.addWidget(self.lbl_variation_note)

        folder_row = QHBoxLayout()
        self.btn_excel_folder = make_btn("\U0001F4C2 Folder Audio")
        self.btn_excel_folder.clicked.connect(self._pick_audio_folder)
        folder_row.addWidget(self.btn_excel_folder)
        self.lbl_excel_folder = QLabel("Belum dipilih")
        self.lbl_excel_folder.setStyleSheet("font-size: 10px; color: #9CA3AF;")
        folder_row.addWidget(self.lbl_excel_folder, 1)
        self.lbl_match_status = QLabel("")
        self.lbl_match_status.setStyleSheet("font-size: 9px; padding: 2px 6px; border-radius: 4px; font-weight: 600;")
        folder_row.addWidget(self.lbl_match_status)
        excel_var_layout.addLayout(folder_row)

        self.btn_apply_variation = make_btn("\u2714 Terapkan Urutan", "#0D9488", "white")
        self.btn_apply_variation.clicked.connect(self._apply_variation_order)
        excel_var_layout.addWidget(self.btn_apply_variation)
        self.excel_variation_widget.setVisible(False)
        excel_layout.addWidget(self.excel_variation_widget)
        excel_layout.addStretch()
        self.excel_audio_panel.setVisible(False)
        l_audio.addWidget(self.excel_audio_panel, 1)

        c_audio.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        col.addWidget(c_audio, 1)

        # Duration card
        c_dur, l_dur = make_card()
        l_dur.addWidget(make_card_title("Mode Durasi"))
        self.radio_audio = QRadioButton("Sesuai total durasi audio")
        self.radio_manual = QRadioButton("Durasi manual")
        self.radio_loop = QRadioButton("Ulangi list")
        self.radio_audio.setChecked(True)
        self.dur_group = QButtonGroup()
        self.dur_group.addButton(self.radio_audio)
        self.dur_group.addButton(self.radio_manual)
        self.dur_group.addButton(self.radio_loop)
        l_dur.addWidget(self.radio_audio)

        row_manual = QHBoxLayout()
        row_manual.setSpacing(4)
        row_manual.addWidget(self.radio_manual)
        self.entry_jam = QLineEdit("00"); self.entry_jam.setFixedWidth(44); self.entry_jam.setAlignment(Qt.AlignCenter)
        self.entry_menit = QLineEdit("00"); self.entry_menit.setFixedWidth(44); self.entry_menit.setAlignment(Qt.AlignCenter)
        self.entry_detik = QLineEdit("00"); self.entry_detik.setFixedWidth(44); self.entry_detik.setAlignment(Qt.AlignCenter)
        row_manual.addWidget(self.entry_jam); row_manual.addWidget(QLabel(":"))
        row_manual.addWidget(self.entry_menit); row_manual.addWidget(QLabel(":"))
        row_manual.addWidget(self.entry_detik); row_manual.addStretch()
        l_dur.addLayout(row_manual)

        row_loop = QHBoxLayout()
        row_loop.setSpacing(4)
        row_loop.addWidget(self.radio_loop)
        self.entry_loop = QLineEdit("2"); self.entry_loop.setFixedWidth(44); self.entry_loop.setAlignment(Qt.AlignCenter)
        row_loop.addWidget(self.entry_loop); row_loop.addWidget(QLabel("\u00d7")); row_loop.addStretch()
        l_dur.addLayout(row_loop)
        col.addWidget(c_dur)


    # ═══ COLUMN 2: RENDER OPTIONS + YOUTUBE DETAIL ═══════════════════════════
    def _build_col2_render_youtube(self, col):
        # Direct layout (no scroll) so column 2 shows fully without slider
        scroll_layout = col

        # Render options card
        c_render, l_render = make_card()
        l_render.addWidget(make_card_title("\u2699\uFE0F Opsi Render"))
        grid = QGridLayout()
        grid.setSpacing(8)

        for c, text in enumerate(["Encoder", "Mode", "FPS"]):
            lbl = QLabel(text)
            lbl.setStyleSheet("font-size: 10px; font-weight: bold; color: #6B7280;")
            grid.addWidget(lbl, 0, c)

        self.combo_encoder = ModernComboBox()
        self.combo_encoder.addItems([n for n, _ in self.daftar_encoder])
        self.combo_mode = ModernComboBox()
        self.combo_mode.addItems(["Super Cepat", "Cepat", "Normal", "Lambat"])
        self.combo_mode.setCurrentIndex(2)
        self.combo_fps = ModernComboBox()
        self.combo_fps.addItems(["10 FPS", "15 FPS", "24 FPS", "30 FPS"])
        self.combo_fps.setCurrentIndex(2)
        grid.addWidget(self.combo_encoder, 1, 0)
        grid.addWidget(self.combo_mode, 1, 1)
        grid.addWidget(self.combo_fps, 1, 2)

        lbl_b = QLabel("Audio Bitrate")
        lbl_b.setStyleSheet("font-size: 10px; font-weight: bold; color: #6B7280;")
        grid.addWidget(lbl_b, 2, 0)
        lbl_cf = QLabel("Crossfade")
        lbl_cf.setStyleSheet("font-size: 10px; font-weight: bold; color: #6B7280;")
        grid.addWidget(lbl_cf, 2, 1, 1, 2)

        self.combo_bitrate = ModernComboBox()
        self.combo_bitrate.addItems(["192 kbps", "256 kbps", "320 kbps"])
        self.combo_bitrate.setCurrentIndex(1)
        grid.addWidget(self.combo_bitrate, 3, 0)
        self.chk_crossfade = QCheckBox("Aktif")
        self.chk_crossfade.toggled.connect(lambda c: self.combo_crossfade_dur.setEnabled(c))
        grid.addWidget(self.chk_crossfade, 3, 1)
        self.combo_crossfade_dur = ModernComboBox()
        self.combo_crossfade_dur.addItems(["1 detik", "2 detik", "3 detik", "4 detik", "5 detik"])
        self.combo_crossfade_dur.setCurrentIndex(1)
        self.combo_crossfade_dur.setEnabled(False)
        grid.addWidget(self.combo_crossfade_dur, 3, 2)

        self.chk_overlay = QCheckBox("Overlay Now Playing")
        self.chk_overlay.toggled.connect(lambda c: self.combo_overlay_pos.setEnabled(c))
        self.combo_overlay_pos = ModernComboBox()
        self.combo_overlay_pos.addItems(["Bottom-Left", "Bottom-Center", "Bottom-Right", "Top-Left", "Top-Center", "Top-Right"])
        self.combo_overlay_pos.setCurrentIndex(1)
        self.combo_overlay_pos.setEnabled(False)
        grid.addWidget(self.chk_overlay, 4, 0)
        grid.addWidget(self.combo_overlay_pos, 4, 1, 1, 2)

        l_render.addLayout(grid)
        scroll_layout.addWidget(c_render)

        # YouTube detail card
        c_yt, l_yt = make_card()
        l_yt.setSpacing(6)
        l_yt.addWidget(make_card_title("\U0001F4FA Detail YouTube"))

        # Channel
        ch_row = QHBoxLayout()
        ch_row.setSpacing(6)
        self.combo_channel = ModernComboBox()
        self.combo_channel.setMinimumWidth(120)
        self.combo_channel.currentIndexChanged.connect(self._on_channel_changed)
        ch_row.addWidget(self.combo_channel, 1)
        btn_add_ch = make_btn("+ Login", "#0D9488", "white")
        btn_add_ch.clicked.connect(self._add_channel)
        ch_row.addWidget(btn_add_ch)
        l_yt.addLayout(ch_row)

        # Title
        title_hdr = QHBoxLayout()
        title_hdr.addWidget(QLabel("Judul"))
        self.lbl_title_count = QLabel("0 / 100")
        self.lbl_title_count.setStyleSheet("font-size: 10px; color: #9CA3AF;")
        title_hdr.addStretch()
        title_hdr.addWidget(self.lbl_title_count)
        l_yt.addLayout(title_hdr)
        self.entry_title = QLineEdit()
        self.entry_title.setPlaceholderText("Judul video YouTube...")
        self.entry_title.setMaxLength(100)
        self.entry_title.textChanged.connect(self._on_title_changed)
        l_yt.addWidget(self.entry_title)

        # Description
        desc_hdr = QHBoxLayout()
        desc_hdr.addWidget(QLabel("Deskripsi"))
        self.lbl_desc_count = QLabel("0 / 5000")
        self.lbl_desc_count.setStyleSheet("font-size: 10px; color: #9CA3AF;")
        desc_hdr.addStretch()
        desc_hdr.addWidget(self.lbl_desc_count)
        l_yt.addLayout(desc_hdr)
        self.entry_desc = QPlainTextEdit()
        self.entry_desc.setPlaceholderText("Deskripsi video...")
        self.entry_desc.setMaximumHeight(60)
        self.entry_desc.textChanged.connect(self._on_desc_changed)
        l_yt.addWidget(self.entry_desc)

        # Tags
        tags_hdr = QHBoxLayout()
        tags_hdr.addWidget(QLabel("Tags (koma)"))
        self.lbl_tags_count = QLabel("0 / 500")
        self.lbl_tags_count.setStyleSheet("font-size: 10px; color: #9CA3AF;")
        tags_hdr.addStretch()
        tags_hdr.addWidget(self.lbl_tags_count)
        l_yt.addLayout(tags_hdr)
        self.entry_tags = QLineEdit()
        self.entry_tags.setPlaceholderText("tag1, tag2, tag3...")
        self.entry_tags.textChanged.connect(self._on_tags_changed)
        l_yt.addWidget(self.entry_tags)

        # Category + Playlist + Privacy
        row_cpp = QHBoxLayout()
        cat_col = QVBoxLayout()
        cat_col.addWidget(QLabel("Kategori"))
        self.combo_category = ModernComboBox()
        self.combo_category.addItems(list(YOUTUBE_CATEGORIES.keys()))
        self.combo_category.setCurrentText("Music")
        cat_col.addWidget(self.combo_category)
        row_cpp.addLayout(cat_col)

        pp_left = QVBoxLayout()
        pp_left.addWidget(QLabel("Playlist"))
        self.combo_playlist = ModernComboBox()
        self.combo_playlist.addItem("(Tidak ada)")
        pp_left.addWidget(self.combo_playlist)
        row_cpp.addLayout(pp_left)

        pp_right = QVBoxLayout()
        pp_right.addWidget(QLabel("Privacy"))
        self.combo_privacy = ModernComboBox()
        self.combo_privacy.addItems(["Public", "Unlisted", "Private"])
        self.combo_privacy.setCurrentIndex(2)
        self.combo_privacy.currentTextChanged.connect(self._on_privacy_changed)
        pp_right.addWidget(self.combo_privacy)
        row_cpp.addLayout(pp_right)
        l_yt.addLayout(row_cpp)

        # Schedule
        self.chk_schedule = QCheckBox("Jadwalkan Publish (WIT)")
        self.chk_schedule.toggled.connect(self._toggle_schedule)
        l_yt.addWidget(self.chk_schedule)

        self.schedule_container = QWidget()
        sched_layout = QHBoxLayout(self.schedule_container)
        sched_layout.setContentsMargins(0, 0, 0, 0)
        sched_layout.setSpacing(4)
        self.combo_day = ModernComboBox()
        for d in range(1, 32):
            self.combo_day.addItem(str(d).zfill(2), d)
        self.combo_month = ModernComboBox()
        months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        for i, m in enumerate(months, 1):
            self.combo_month.addItem(m, i)
        self.combo_year = ModernComboBox()
        current_year = datetime.datetime.now().year
        for y in range(current_year, current_year + 3):
            self.combo_year.addItem(str(y), y)
        self.combo_hour = ModernComboBox()
        for h in range(24):
            self.combo_hour.addItem(str(h).zfill(2), h)
        self.combo_minute = ModernComboBox()
        for m in range(0, 60, 5):
            self.combo_minute.addItem(str(m).zfill(2), m)
        for w in [self.combo_day, self.combo_month, self.combo_year, self.combo_hour, self.combo_minute]:
            w.setStyleSheet("QComboBox { border: none; background: transparent; padding: 4px 2px; font-size: 11px; font-weight: 600; min-width: 36px; } QComboBox::drop-down { border: none; width: 14px; } QComboBox::down-arrow { image: none; }")
        sched_layout.addWidget(self.combo_day)
        sched_layout.addWidget(QLabel("/"))
        sched_layout.addWidget(self.combo_month)
        sched_layout.addWidget(QLabel("/"))
        sched_layout.addWidget(self.combo_year)
        sched_layout.addWidget(QLabel(" "))
        sched_layout.addWidget(self.combo_hour)
        sched_layout.addWidget(QLabel(":"))
        sched_layout.addWidget(self.combo_minute)
        lbl_wit = QLabel("WIT")
        lbl_wit.setStyleSheet("background: #ECFDF5; color: #059669; padding: 3px 6px; border-radius: 4px; font-size: 9px; font-weight: bold;")
        sched_layout.addWidget(lbl_wit)
        sched_layout.addStretch()
        self.schedule_container.setEnabled(False)
        l_yt.addWidget(self.schedule_container)

        # Set default schedule to tomorrow WIT
        now_wit = datetime.datetime.now(WIT_TZ) + datetime.timedelta(days=1)
        self.combo_day.setCurrentIndex(now_wit.day - 1)
        self.combo_month.setCurrentIndex(now_wit.month - 1)
        self.combo_hour.setCurrentIndex(now_wit.hour)
        minute_idx = now_wit.minute // 5
        if minute_idx < self.combo_minute.count():
            self.combo_minute.setCurrentIndex(minute_idx)

        # Thumbnail
        thumb_row = QHBoxLayout()
        self.btn_thumb = make_btn("\U0001F5BC Thumbnail")
        self.btn_thumb.clicked.connect(self._pick_thumbnail)
        self.lbl_thumb = QLabel("(opsional)")
        self.lbl_thumb.setStyleSheet("font-size: 10px; color: #9CA3AF;")
        thumb_row.addWidget(self.btn_thumb)
        thumb_row.addWidget(self.lbl_thumb, 1)
        l_yt.addLayout(thumb_row)
        self._thumbnail_path = None

        scroll_layout.addWidget(c_yt, 1)

        # Add to pipeline button
        self.btn_add_pipeline = QPushButton("+ Tambah ke Antrian Pipeline")
        self.btn_add_pipeline.setCursor(Qt.PointingHandCursor)
        self.btn_add_pipeline.setStyleSheet(
            "QPushButton { background-color: #0D9488; color: white; border: none; border-radius: 10px; padding: 12px; font-size: 12px; font-weight: bold; }"
            "QPushButton:hover { background-color: #0B7C72; }"
        )
        self.btn_add_pipeline.clicked.connect(self._add_to_pipeline)
        col.addWidget(self.btn_add_pipeline)


    # ═══ COLUMN 3: TRACKLIST / TIMESTAMP ═════════════════════════════════════
    def _build_col3_tracklist(self, col):
        c_track, l_track = make_card()
        hdr = QHBoxLayout()
        hdr.addWidget(make_card_title("\U0001F4CB Tracklist / Timestamp"))
        hdr.addStretch()
        self.btn_copy = make_btn("Copy")
        self.btn_copy.clicked.connect(self._copy_tracklist)
        hdr.addWidget(self.btn_copy)
        l_track.addLayout(hdr)

        self.tracklist_area = QTextEdit()
        self.tracklist_area.setReadOnly(True)
        self.tracklist_area.setStyleSheet(
            "background-color: #F9FAFB; color: #374151; border: 1px solid #E5E7EB; border-radius: 10px; padding: 10px;"
            "font-family: 'Cascadia Code', 'Consolas', monospace; font-size: 11px;"
        )
        l_track.addWidget(self.tracklist_area, 1)

        # Estimasi
        self.lbl_estimate = QLabel("Tambahkan audio untuk estimasi")
        self.lbl_estimate.setStyleSheet("background: #F0FDFA; color: #065F46; padding: 6px 10px; border-radius: 6px; font-size: 10px; font-weight: 600;")
        l_track.addWidget(self.lbl_estimate)

        c_track.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        col.addWidget(c_track, 1)

        # Connect signals for auto-update
        self.radio_audio.toggled.connect(self._update_tracklist)
        self.radio_loop.toggled.connect(self._update_tracklist)
        self.entry_loop.textChanged.connect(self._update_tracklist)
        self.chk_crossfade.toggled.connect(self._update_tracklist)
        self.combo_crossfade_dur.currentIndexChanged.connect(self._update_tracklist)
        self.combo_mode.currentIndexChanged.connect(self._update_size_estimate)
        self.combo_encoder.currentIndexChanged.connect(self._update_size_estimate)
        self.combo_fps.currentIndexChanged.connect(self._update_size_estimate)
        self.combo_bitrate.currentIndexChanged.connect(self._update_size_estimate)

    # ═══ COLUMN 4: PIPELINE QUEUE + LOG ═══════════════════════════════════════
    def _build_col4_pipeline(self, col):
        # Slot indicators
        c_slots, l_slots = make_card()
        l_slots.addWidget(make_card_title("\U0001F504 Pipeline Slots"))
        slot_row = QHBoxLayout()
        self.lbl_encode_slot = QLabel("\u25CF Encode: idle")
        self.lbl_encode_slot.setStyleSheet("font-size: 10px; color: #6B7280;")
        self.lbl_upload_slot = QLabel("\u25CF Upload: idle")
        self.lbl_upload_slot.setStyleSheet("font-size: 10px; color: #6B7280;")
        slot_row.addWidget(self.lbl_encode_slot)
        slot_row.addWidget(self.lbl_upload_slot)
        l_slots.addLayout(slot_row)
        col.addWidget(c_slots)

        # Queue
        c_queue, l_queue = make_card()
        l_queue.addWidget(make_card_title("Antrian Pipeline"))
        self.queue_scroll = QScrollArea()
        self.queue_scroll.setWidgetResizable(True)
        self.queue_scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        self.queue_container = QWidget()
        self.queue_container.setStyleSheet("background: transparent;")
        self.queue_list_layout = QVBoxLayout(self.queue_container)
        self.queue_list_layout.setContentsMargins(0, 0, 0, 0)
        self.queue_list_layout.setSpacing(6)
        self.queue_list_layout.addStretch()
        self.queue_scroll.setWidget(self.queue_container)
        l_queue.addWidget(self.queue_scroll, 1)

        self.btn_start_pipeline = QPushButton("\u25b6 Mulai Pipeline")
        self.btn_start_pipeline.setCursor(Qt.PointingHandCursor)
        self.btn_start_pipeline.setStyleSheet(
            "QPushButton { background-color: qlineargradient(x1:0,y1:0,x2:1,y2:1,stop:0 #0D9488,stop:1 #059669);"
            "color: white; border: none; border-radius: 10px; padding: 12px; font-size: 12px; font-weight: bold; }"
            "QPushButton:hover { background-color: qlineargradient(x1:0,y1:0,x2:1,y2:1,stop:0 #0B7C72,stop:1 #047857); }"
        )
        self.btn_start_pipeline.clicked.connect(self._start_pipeline)
        l_queue.addWidget(self.btn_start_pipeline)

        c_queue.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        col.addWidget(c_queue, 1)

        # Log
        c_log, l_log = make_card()
        l_log.addWidget(make_card_title("Log"))
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setStyleSheet(
            "background-color: #1F2937; color: #5EEAD4; border: none; border-radius: 10px; padding: 10px;"
            "font-family: 'Cascadia Code', 'Consolas', monospace; font-size: 10px;"
        )
        l_log.addWidget(self.log_area, 1)
        c_log.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        col.addWidget(c_log)


    # ═══════════════════════════════════════════════════════════════════════════
    # MEDIA ACTIONS (Column 1)
    # ═══════════════════════════════════════════════════════════════════════════

    def _pilih_video(self):
        jalur, _ = QFileDialog.getOpenFileName(self, "Pilih File Video", "", "Video (*.mp4 *.avi *.mov *.mkv)")
        if jalur:
            self.jalur_video = jalur
            self.lbl_video.setText(os.path.basename(jalur))
            self.lbl_video.setStyleSheet("font-size: 11px; font-weight: 500; color: #1a1a2e;")
            self._update_video_preview(jalur)
            self._log(f"Video: {os.path.basename(jalur)}")

    def _update_video_preview(self, path):
        preview_file = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
        preview_path = preview_file.name
        preview_file.close()
        try:
            ffmpeg_exe = imageio_ffmpeg.get_ffmpeg_exe()
            subprocess.run(
                [ffmpeg_exe, "-y", "-ss", "00:00:01", "-i", path, "-frames:v", "1", "-q:v", "3", preview_path],
                capture_output=True, timeout=20, creationflags=subprocess.CREATE_NO_WINDOW
            )
            pixmap = QPixmap(preview_path)
            if not pixmap.isNull():
                target_w = self.video_preview.width()
                target_h = self.video_preview.height()
                scaled = pixmap.scaled(target_w, target_h, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
                crop_x = max(0, (scaled.width() - target_w) // 2)
                crop_y = max(0, (scaled.height() - target_h) // 2)
                cropped = scaled.copy(crop_x, crop_y, target_w, target_h)
                rounded = QPixmap(target_w, target_h)
                rounded.fill(Qt.transparent)
                painter = QPainter(rounded)
                painter.setRenderHint(QPainter.Antialiasing)
                clip_path = QPainterPath()
                clip_path.addRoundedRect(QRectF(0, 0, target_w, target_h), 10, 10)
                painter.setClipPath(clip_path)
                painter.drawPixmap(0, 0, cropped)
                painter.end()
                self.video_preview.setPixmap(rounded)
                self.video_preview.setText("")
                self.video_preview.setStyleSheet("QLabel { border: none; border-radius: 10px; }")
        except Exception:
            pass
        finally:
            try:
                os.remove(preview_path)
            except Exception:
                pass

    def _tambah_audio(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Pilih File Audio", "", "Audio (*.mp3 *.wav *.m4a *.aac *.flac)")
        added = 0
        for f in files:
            if f not in self.daftar_audio:
                self.daftar_audio.append(f)
                try:
                    clip = AudioFileClip(f)
                    self.audio_durations.append(clip.duration)
                    clip.close()
                except Exception:
                    self.audio_durations.append(0.0)
                added += 1
        if added:
            self._refresh_audio_list()
            self._log(f"Ditambahkan {added} file audio.")

    def _hapus_audio(self):
        if not self.daftar_audio:
            return
        indices_to_remove = []
        for i in range(self.audio_list_widget.count()):
            widget = self.audio_list_widget.get_widget(i)
            if widget and widget.checkbox.isChecked():
                indices_to_remove.append(i)
        if not indices_to_remove:
            indices_to_remove = [len(self.daftar_audio) - 1]
        for idx in sorted(indices_to_remove, reverse=True):
            if 0 <= idx < len(self.daftar_audio):
                del self.daftar_audio[idx]
                del self.audio_durations[idx]
        self._refresh_audio_list()

    def _acak_audio(self):
        if len(self.daftar_audio) > 1:
            combined = list(zip(self.daftar_audio, self.audio_durations))
            random.shuffle(combined)
            self.daftar_audio, self.audio_durations = [list(x) for x in zip(*combined)]
            self._refresh_audio_list()
            self._log("Urutan diacak.")

    def _on_audio_reordered(self):
        self._update_tracklist()

    def _refresh_audio_list(self):
        self.audio_list_widget.clear()
        for i, path in enumerate(self.daftar_audio):
            dur = self.audio_durations[i] if i < len(self.audio_durations) else 0
            dur_str = f"{int(dur//60)}:{int(dur%60):02d}"
            filename = os.path.basename(path)
            widget = AudioItemWidget(i, filename, dur_str)
            self.audio_list_widget.add_item(widget, i)
        self._update_tracklist()

    # ─── Excel Import ────────────────────────────────────────────────────────
    def _set_audio_mode(self, mode):
        self._audio_mode = mode
        self.btn_mode_manual.setChecked(mode == 'manual')
        self.btn_mode_excel.setChecked(mode == 'excel')
        self.manual_audio_panel.setVisible(mode == 'manual')
        self.excel_audio_panel.setVisible(mode == 'excel')

    def _import_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "Error", "openpyxl tidak terinstall.\nJalankan: pip install openpyxl")
            return
        filepath, _ = QFileDialog.getOpenFileName(self, "Import Excel", "", "Excel Files (*.xlsx *.xls)")
        if not filepath:
            return
        self._excel_filepath = filepath
        self._log(f"Import Excel: {os.path.basename(filepath)}")
        result = _parse_excel_variations(filepath)
        if result['error']:
            QMessageBox.warning(self, "Error", result['error'])
            return

        self._excel_data = result
        try:
            self._excel_seo_data, self._excel_seo_ordered = _parse_visual_seo(filepath)
        except Exception as e:
            self._excel_seo_data, self._excel_seo_ordered = {}, []
            self._log(f"SEO parse error: {e}")
        self._log(f"SEO data loaded: {len(self._excel_seo_data)} entries, keys: {list(self._excel_seo_data.keys())[:6]}...")
        if result.get('var_num_map'):
            self._log(f"Var num map: {result['var_num_map']}")

        self.excel_import_zone.setStyleSheet(
            "QPushButton { background: #D1FAE5; border: 2px solid #10B981; border-radius: 10px; padding: 10px; font-size: 11px; font-weight: 600; color: #059669; }"
        )
        num_v = len(result['variation_names'])
        self.excel_import_zone.setText(f"\u2705 {os.path.basename(filepath)} ({num_v} variasi)")

        self.combo_variation.blockSignals(True)
        self.combo_variation.clear()
        for name in result['variation_names']:
            self.combo_variation.addItem(name)
        self.combo_variation.blockSignals(False)
        self.excel_variation_widget.setVisible(True)

        if result['variation_names']:
            self._on_variation_changed(0)

        if self._excel_folder:
            self._do_matching()

    def _on_variation_changed(self, index):
        if not self._excel_data or index < 0:
            return
        var_names = self._excel_data['variation_names']
        if index >= len(var_names):
            return
        var_name = var_names[index]
        note = self._excel_data['variations'][var_name].get('note', '')
        self.lbl_variation_note.setText(note if note else "")

        # Auto-fill YouTube from SEO data
        # Strategy: try multiple lookup keys in order of specificity
        seo = None

        # 1. Direct name match
        if var_name in self._excel_seo_data:
            seo = self._excel_seo_data[var_name]

        # 2. Try variation number from var_num_map
        if not seo:
            var_num_map = self._excel_data.get('var_num_map', {})
            var_num = var_num_map.get(var_name)
            if var_num and var_num in self._excel_seo_data:
                seo = self._excel_seo_data[var_num]

        # 3. Try index+1 as string (positional: variation 1, 2, 3...)
        if not seo:
            idx_key = str(index + 1)
            if idx_key in self._excel_seo_data:
                seo = self._excel_seo_data[idx_key]

        # 4. Positional fallback: use the Nth entry from ordered SEO list
        if not seo:
            seo_ordered = self._excel_seo_ordered
            if index < len(seo_ordered):
                seo = seo_ordered[index]

        if seo:
            if seo.get('title'):
                self.entry_title.setText(seo['title'][:100])
            if seo.get('description'):
                self.entry_desc.setPlainText(seo['description'][:5000])
            if seo.get('tags'):
                self.entry_tags.setText(seo['tags'][:500])
            self._log(f"SEO auto-fill: {seo.get('title', '')[:50]}")
        else:
            self._log(f"SEO data tidak ditemukan untuk variasi: {var_name}")

    def _pick_audio_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Pilih Folder Audio")
        if not folder:
            return
        self._excel_folder = folder
        display = folder if len(folder) <= 30 else "..." + folder[-27:]
        self.lbl_excel_folder.setText(display)
        self.lbl_excel_folder.setToolTip(folder)
        self._do_matching()

    def _do_matching(self):
        if not self._excel_data or not self._excel_folder:
            return
        audio_files = _scan_audio_folder(self._excel_folder)
        if not audio_files:
            self.lbl_match_status.setText("0 file")
            self.lbl_match_status.setStyleSheet("font-size: 9px; padding: 2px 6px; border-radius: 4px; font-weight: 600; background: #FEF3C7; color: #D97706;")
            return
        self._excel_matches = _match_all_titles(self._excel_data, audio_files)
        matched = sum(1 for v in self._excel_matches.values() if v is not None)
        total = len(self._excel_matches)
        if matched == total:
            self.lbl_match_status.setText(f"{matched}/{total} \u2714")
            self.lbl_match_status.setStyleSheet("font-size: 9px; padding: 2px 6px; border-radius: 4px; font-weight: 600; background: #D1FAE5; color: #059669;")
        else:
            self.lbl_match_status.setText(f"{matched}/{total}")
            self.lbl_match_status.setStyleSheet("font-size: 9px; padding: 2px 6px; border-radius: 4px; font-weight: 600; background: #FEF3C7; color: #D97706;")

    def _apply_variation_order(self):
        if not self._excel_data or not self._excel_folder or not self._excel_matches:
            QMessageBox.warning(self, "Error", "Import Excel dan pilih folder audio terlebih dahulu!")
            return
        var_index = self.combo_variation.currentIndex()
        var_names = self._excel_data['variation_names']
        if var_index < 0 or var_index >= len(var_names):
            return
        var_name = var_names[var_index]
        tracks = self._excel_data['variations'][var_name]['tracks']

        new_audio = []
        new_durations = []
        for track in tracks:
            title = track['title']
            file_path = self._excel_matches.get(title)
            if file_path and os.path.exists(file_path):
                new_audio.append(file_path)
                try:
                    clip = AudioFileClip(file_path)
                    new_durations.append(clip.duration)
                    clip.close()
                except Exception:
                    new_durations.append(0.0)

        if not new_audio:
            QMessageBox.warning(self, "Error", "Tidak ada file audio yang cocok!")
            return

        self.daftar_audio = new_audio
        self.audio_durations = new_durations
        self._set_audio_mode('manual')
        self._refresh_audio_list()
        self._log(f"Variasi '{var_name}' diterapkan: {len(new_audio)} lagu")


    # ═══════════════════════════════════════════════════════════════════════════
    # TRACKLIST & ESTIMATION (Column 3)
    # ═══════════════════════════════════════════════════════════════════════════

    @staticmethod
    def _strip_track_number(name):
        stripped = re.sub(r'^\d+[\.\-_\s]+\s*', '', name)
        return stripped if stripped else name

    def _update_tracklist(self):
        if not self.daftar_audio:
            self.tracklist_area.clear()
            self._update_size_estimate()
            return

        loop_count = 1
        if self.radio_loop.isChecked():
            try:
                loop_count = max(1, int(self.entry_loop.text()))
            except ValueError:
                loop_count = 1

        crossfade_dur = 0
        if self.chk_crossfade.isChecked():
            try:
                crossfade_dur = int(self.combo_crossfade_dur.currentText().split()[0])
            except (ValueError, IndexError):
                crossfade_dur = 0

        total_tracks = len(self.daftar_audio) * loop_count
        track_num = 0
        lines = []
        waktu = 0.0

        for loop_i in range(loop_count):
            if loop_count > 1 and loop_i > 0:
                lines.append(f"\n\u2500\u2500 Putaran {loop_i + 1} \u2500\u2500")
            for i, path in enumerate(self.daftar_audio):
                track_num += 1
                ts = self._format_waktu(waktu)
                nama = os.path.splitext(os.path.basename(path))[0]
                nama = self._strip_track_number(nama)
                lines.append(f"{ts}  {nama}")
                dur = self.audio_durations[i] if i < len(self.audio_durations) else 0
                effective_dur = dur - crossfade_dur if (track_num < total_tracks and crossfade_dur > 0) else dur
                waktu += max(0, effective_dur)

        self.tracklist_area.setPlainText("\n".join(lines))
        self._update_size_estimate()

    def _estimate_duration(self):
        if not self.daftar_audio:
            return 0.0
        if self.radio_manual.isChecked():
            try:
                h = int(self.entry_jam.text() or 0)
                m = int(self.entry_menit.text() or 0)
                s = int(self.entry_detik.text() or 0)
                return max(0, h * 3600 + m * 60 + s)
            except ValueError:
                return 0.0
        loop_count = 1
        if self.radio_loop.isChecked():
            try:
                loop_count = max(1, int(self.entry_loop.text() or 1))
            except ValueError:
                loop_count = 1
        total_tracks = len(self.daftar_audio) * loop_count
        total_dur = sum(self.audio_durations) * loop_count
        if self.chk_crossfade.isChecked() and total_tracks > 1:
            try:
                crossfade_dur = int(self.combo_crossfade_dur.currentText().split()[0])
            except (ValueError, IndexError):
                crossfade_dur = 0
            total_dur -= crossfade_dur * (total_tracks - 1)
        return max(0.0, total_dur)

    def _compression_preset(self, mode, encoder, fps=None):
        presets = {
            "Super Cepat": {"bitrate": 1800, "maxrate": 2500, "bufsize": 5000, "cq": "30", "nvenc_preset": "p1", "x264_preset": "veryfast"},
            "Cepat": {"bitrate": 2500, "maxrate": 3500, "bufsize": 7000, "cq": "28", "nvenc_preset": "p2", "x264_preset": "veryfast"},
            "Normal": {"bitrate": 3500, "maxrate": 5000, "bufsize": 10000, "cq": "26", "nvenc_preset": "p4", "x264_preset": "medium"},
            "Lambat": {"bitrate": 4500, "maxrate": 6500, "bufsize": 13000, "cq": "24", "nvenc_preset": "p5", "x264_preset": "slow"},
        }
        preset = dict(presets.get(mode, presets["Normal"]))
        if encoder == "libx264":
            preset["bitrate"] = int(preset["bitrate"] * 0.85)
            preset["maxrate"] = int(preset["maxrate"] * 0.9)
            preset["bufsize"] = int(preset["bufsize"] * 0.9)
        elif encoder in ("h264_amf", "h264_qsv"):
            preset["bitrate"] = int(preset["bitrate"] * 1.05)
            preset["maxrate"] = int(preset["maxrate"] * 1.05)
            preset["bufsize"] = int(preset["bufsize"] * 1.05)
        if fps:
            try:
                fps_factor = max(0.6, int(fps) / 24)
            except (TypeError, ValueError):
                fps_factor = 1.0
            preset["bitrate"] = int(preset["bitrate"] * fps_factor)
            preset["maxrate"] = int(preset["maxrate"] * fps_factor)
            preset["bufsize"] = int(preset["bufsize"] * fps_factor)
        return preset

    def _update_size_estimate(self):
        total_dur = self._estimate_duration()
        if total_dur <= 0:
            self.lbl_estimate.setText("Tambahkan audio untuk estimasi")
            return
        encoder = self.daftar_encoder[self.combo_encoder.currentIndex()][1]
        mode = self.combo_mode.currentText()
        fps = self.combo_fps.currentText().split()[0]
        preset = self._compression_preset(mode, encoder, fps)
        video_kbps = preset['bitrate']
        audio_kbps = int(self.combo_bitrate.currentText().split()[0])
        total_kbps = video_kbps + audio_kbps
        size_bytes = total_dur * total_kbps * 1000 / 8
        size_str = f"{size_bytes / (1024*1024):.0f} MB" if size_bytes < 1024**3 else f"{size_bytes / (1024**3):.2f} GB"
        dur_str = self._format_waktu(total_dur)
        self.lbl_estimate.setText(f"\u00b1 {size_str}  \u2022  {dur_str}  \u2022  {video_kbps}k video + {audio_kbps}k audio")

    def _copy_tracklist(self):
        text = self.tracklist_area.toPlainText()
        if text:
            QApplication.clipboard().setText(text)
            self._log("Tracklist disalin ke clipboard.")

    def _get_tracklist_text(self):
        """Get tracklist text for injection into YouTube description."""
        return self.tracklist_area.toPlainText()


    # ═══════════════════════════════════════════════════════════════════════════
    # YOUTUBE ACTIONS (Column 2)
    # ═══════════════════════════════════════════════════════════════════════════

    def _add_channel(self):
        if not CLIENT_SECRET.exists():
            QMessageBox.warning(self, "Error", f"client_secret.json tidak ditemukan!\nLetakkan di:\n{CLIENT_SECRET}")
            return
        from PySide6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self, "Nama Channel", "Masukkan nama/alias untuk channel:")
        if not ok or not name.strip():
            return
        name = name.strip()
        self._log(f"Authenticating '{name}'...")
        QApplication.processEvents()

        def do_auth():
            youtube, err = YouTubeAuth.authenticate(name)
            if err:
                self.signals.log.emit(f"Auth gagal: {err}")
                return
            ch_title, ch_id, thumb_url = YouTubeAuth.get_channel_info(youtube)
            if ch_title:
                self.signals.log.emit(f"Login berhasil: {ch_title}")
                self.channel_info[name] = {"title": ch_title, "id": ch_id}
            self.channels[name] = youtube
            self.combo_channel.addItem(name)
            self.combo_channel.setCurrentText(name)

        threading.Thread(target=do_auth, daemon=True).start()

    def _on_channel_changed(self, index):
        name = self.combo_channel.currentText()
        if not name:
            return
        if name not in self.channels:
            youtube, err = YouTubeAuth.get_service(name)
            if youtube:
                self.channels[name] = youtube
                if name not in self.channel_info:
                    ch_title, ch_id, _ = YouTubeAuth.get_channel_info(youtube)
                    if ch_title:
                        self.channel_info[name] = {"title": ch_title, "id": ch_id}
            else:
                self._log(f"Channel '{name}': {err}")
                return
        self._refresh_playlists()

    def _refresh_playlists(self):
        self.combo_playlist.clear()
        self.combo_playlist.addItem("(Tidak ada)")
        name = self.combo_channel.currentText()
        if name and name in self.channels:
            self.playlists = YouTubeAuth.get_playlists(self.channels[name])
            for pl in self.playlists:
                self.combo_playlist.addItem(pl["title"])

    def _pick_thumbnail(self):
        path, _ = QFileDialog.getOpenFileName(self, "Pilih Thumbnail", "", "Images (*.jpg *.jpeg *.png)")
        if path:
            self._thumbnail_path = self._resize_thumbnail(path)
            self.lbl_thumb.setText(os.path.basename(path))
            self.lbl_thumb.setStyleSheet("font-size: 10px; color: #059669; font-weight: 500;")

    def _resize_thumbnail(self, path):
        """Resize thumbnail to 1280x720 for YouTube upload."""
        pixmap = QPixmap(path)
        if pixmap.isNull():
            return path
        target_w, target_h = 1280, 720
        source_ratio = pixmap.width() / pixmap.height()
        target_ratio = target_w / target_h
        if source_ratio > target_ratio:
            crop_h = pixmap.height()
            crop_w = int(crop_h * target_ratio)
            crop_x = (pixmap.width() - crop_w) // 2
            crop_y = 0
        else:
            crop_w = pixmap.width()
            crop_h = int(crop_w / target_ratio)
            crop_x = 0
            crop_y = (pixmap.height() - crop_h) // 2
        cropped = pixmap.copy(crop_x, crop_y, crop_w, crop_h)
        resized = cropped.scaled(target_w, target_h, Qt.IgnoreAspectRatio, Qt.SmoothTransformation)
        THUMBNAIL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(THUMBNAIL_CACHE_DIR / f"thumb_{stamp}_1280x720.jpg")
        resized.save(output_path, "JPG", 90)
        return output_path

    def _on_title_changed(self, text):
        count = len(text)
        self.lbl_title_count.setText(f"{count} / 100")
        color = "#DC2626" if count > 100 else "#D97706" if count > 80 else "#9CA3AF"
        self.lbl_title_count.setStyleSheet(f"font-size: 10px; color: {color};")

    def _on_desc_changed(self):
        count = len(self.entry_desc.toPlainText())
        self.lbl_desc_count.setText(f"{count} / 5000")
        color = "#DC2626" if count > 5000 else "#D97706" if count > 4500 else "#9CA3AF"
        self.lbl_desc_count.setStyleSheet(f"font-size: 10px; color: {color};")

    def _on_tags_changed(self, text):
        count = len(text)
        self.lbl_tags_count.setText(f"{count} / 500")
        color = "#DC2626" if count > 500 else "#D97706" if count > 400 else "#9CA3AF"
        self.lbl_tags_count.setStyleSheet(f"font-size: 10px; color: {color};")

    def _toggle_schedule(self, checked):
        self.schedule_container.setEnabled(checked)
        if checked:
            self.combo_privacy.setCurrentText("Private")

    def _on_privacy_changed(self, text):
        if text != "Private" and self.chk_schedule.isChecked():
            self.chk_schedule.setChecked(False)


    # ═══════════════════════════════════════════════════════════════════════════
    # PIPELINE QUEUE MANAGEMENT (Column 4)
    # ═══════════════════════════════════════════════════════════════════════════

    def _add_to_pipeline(self):
        """Validate all fields and add a unified encode+upload task to pipeline."""
        # Validate video
        if not self.jalur_video:
            QMessageBox.warning(self, "Error", "Pilih video utama terlebih dahulu!")
            return
        if not self.daftar_audio:
            QMessageBox.warning(self, "Error", "Tambahkan minimal satu file audio!")
            return

        # Validate YouTube
        yt_title = self.entry_title.text().strip()
        if not yt_title:
            QMessageBox.warning(self, "Error", "Judul YouTube wajib diisi!")
            return
        if len(yt_title) > 100:
            QMessageBox.warning(self, "Error", "Judul melebihi 100 karakter!")
            return
        if len(self.entry_desc.toPlainText()) > 5000:
            QMessageBox.warning(self, "Error", "Deskripsi melebihi 5000 karakter!")
            return
        if len(self.entry_tags.text()) > 500:
            QMessageBox.warning(self, "Error", "Tags melebihi 500 karakter!")
            return
        channel_name = self.combo_channel.currentText()
        if not channel_name or channel_name not in self.channels:
            QMessageBox.warning(self, "Error", "Login ke channel YouTube terlebih dahulu!")
            return

        # Build output path
        jalur_simpan, _ = QFileDialog.getSaveFileName(self, "Simpan Output Video", "", "MP4 Video (*.mp4)")
        if not jalur_simpan:
            return

        # Gather encode params
        idx = self.combo_encoder.currentIndex()
        encoder_codec = self.daftar_encoder[idx][1]
        mode_dur = "audio"
        durasi_manual = None
        jumlah_loop = None
        if self.radio_manual.isChecked():
            mode_dur = "manual"
            try:
                h = int(self.entry_jam.text())
                m = int(self.entry_menit.text())
                s = int(self.entry_detik.text())
                durasi_manual = h * 3600 + m * 60 + s
                if durasi_manual <= 0:
                    raise ValueError
            except ValueError:
                QMessageBox.warning(self, "Error", "Format durasi tidak valid!")
                return
        elif self.radio_loop.isChecked():
            mode_dur = "loop"
            try:
                jumlah_loop = max(1, int(self.entry_loop.text()))
            except ValueError:
                QMessageBox.warning(self, "Error", "Jumlah pengulangan tidak valid!")
                return

        crossfade_enabled = self.chk_crossfade.isChecked()
        crossfade_dur = int(self.combo_crossfade_dur.currentText().split()[0]) if crossfade_enabled else 0
        audio_bitrate = self.combo_bitrate.currentText().split()[0] + "k"

        # YouTube params
        privacy = self.combo_privacy.currentText().lower()
        category_name = self.combo_category.currentText() if hasattr(self, 'combo_category') else "Music"
        category_id = YOUTUBE_CATEGORIES.get(category_name, "10")
        tags_raw = self.entry_tags.text().strip()
        tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

        scheduled_time = None
        if self.chk_schedule.isChecked():
            day = self.combo_day.currentData()
            month = self.combo_month.currentData()
            year = self.combo_year.currentData()
            hour = self.combo_hour.currentData()
            minute = self.combo_minute.currentData()
            try:
                local_dt = datetime.datetime(year, month, day, hour, minute, tzinfo=WIT_TZ)
                utc_dt = local_dt.astimezone(ZoneInfo("UTC"))
                scheduled_time = utc_dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")
            except ValueError:
                QMessageBox.warning(self, "Error", "Tanggal tidak valid!")
                return

        playlist_id = None
        pl_idx = self.combo_playlist.currentIndex()
        if pl_idx > 0 and (pl_idx - 1) < len(self.playlists):
            playlist_id = self.playlists[pl_idx - 1]["id"]

        # Inject timestamp into description
        description = self.entry_desc.toPlainText()
        tracklist_text = self._get_tracklist_text()
        if tracklist_text:
            description = f"{description}\n\n\U0001F3B5 Tracklist:\n{tracklist_text}"
            if len(description) > 5000:
                description = description[:5000]

        # Build unified task
        task = {
            # Encode params
            'video': self.jalur_video,
            'audio': list(self.daftar_audio),
            'audio_durations': list(self.audio_durations),
            'output_path': jalur_simpan,
            'kompresi': self.combo_mode.currentText(),
            'fps': self.combo_fps.currentText().split()[0],
            'encoder': encoder_codec,
            'mode_durasi': mode_dur,
            'durasi_manual': durasi_manual,
            'jumlah_loop': jumlah_loop,
            'overlay': self.chk_overlay.isChecked(),
            'overlay_pos': self.combo_overlay_pos.currentText() if self.chk_overlay.isChecked() else None,
            'audio_bitrate': audio_bitrate,
            'crossfade': crossfade_enabled,
            'crossfade_dur': crossfade_dur,
            # Upload params
            'channel': channel_name,
            'yt_title': yt_title,
            'yt_description': description,
            'yt_tags': tags,
            'yt_category_id': category_id,
            'yt_privacy': privacy,
            'yt_scheduled_time': scheduled_time,
            'yt_thumbnail_path': self._thumbnail_path,
            'yt_playlist_id': playlist_id,
            # Status
            'status': 'waiting',  # waiting → encoding → uploading → done / failed
            'video_path': jalur_simpan,  # used by uploader after encode
        }

        self.pipeline_queue.append(task)
        self._add_pipeline_widget(task)
        self._log(f"Pipeline +: {yt_title} [{encoder_codec}]")

    def _add_pipeline_widget(self, task):
        task_index = len(self.pipeline_widgets)
        row = QHBoxLayout()
        row.setContentsMargins(10, 8, 10, 8)
        row.setSpacing(10)

        # Number
        lbl_num = QLabel(str(task_index + 1))
        lbl_num.setFixedSize(22, 22)
        lbl_num.setAlignment(Qt.AlignCenter)
        lbl_num.setStyleSheet("background-color: #E5E7EB; color: #6B7280; border-radius: 11px; font-size: 10px; font-weight: bold;")
        row.addWidget(lbl_num)

        # Details
        details = QVBoxLayout()
        details.setSpacing(2)
        title_text = task['yt_title']
        if len(title_text) > 40:
            title_text = title_text[:37] + "..."
        lbl_name = QLabel(title_text)
        lbl_name.setStyleSheet("font-size: 11px; font-weight: bold; color: #1a1a2e;")
        lbl_name.setToolTip(task['yt_title'])

        # Pipeline status: Encode → Upload
        self.status_widget = QHBoxLayout()
        lbl_encode_status = QLabel("\u23F3 Encode")
        lbl_encode_status.setStyleSheet("font-size: 9px; padding: 2px 5px; border-radius: 8px; background: #F3F4F6; color: #9CA3AF;")
        lbl_arrow = QLabel("\u2192")
        lbl_arrow.setStyleSheet("font-size: 9px; color: #D1D5DB;")
        lbl_upload_status = QLabel("\u23F3 Upload")
        lbl_upload_status.setStyleSheet("font-size: 9px; padding: 2px 5px; border-radius: 8px; background: #F3F4F6; color: #9CA3AF;")

        status_row = QHBoxLayout()
        status_row.setSpacing(4)
        status_row.addWidget(lbl_encode_status)
        status_row.addWidget(lbl_arrow)
        status_row.addWidget(lbl_upload_status)
        status_row.addStretch()

        details.addWidget(lbl_name)
        details.addLayout(status_row)
        row.addLayout(details, 1)

        # Delete button
        btn_delete = QPushButton("\u2715")
        btn_delete.setFixedSize(22, 22)
        btn_delete.setStyleSheet(
            "QPushButton { background: none; border: none; color: #DC2626; font-size: 12px; border-radius: 4px; }"
            "QPushButton:hover { background: #FEE2E2; }"
        )
        btn_delete.clicked.connect(lambda checked, idx=task_index: self._delete_pipeline_task(idx))
        row.addWidget(btn_delete)

        frame = QFrame()
        frame.setStyleSheet("QFrame { background-color: #F9FAFB; border: 1px solid #E5E7EB; border-radius: 8px; }")
        frame.setLayout(row)

        count = self.queue_list_layout.count()
        self.queue_list_layout.insertWidget(count - 1, frame)
        self.pipeline_widgets.append({
            'frame': frame,
            'lbl_num': lbl_num,
            'lbl_encode_status': lbl_encode_status,
            'lbl_upload_status': lbl_upload_status,
            'btn_delete': btn_delete,
        })

    def _delete_pipeline_task(self, task_index):
        if task_index >= len(self.pipeline_queue):
            return
        task = self.pipeline_queue[task_index]
        if task['status'] in ('encoding', 'uploading'):
            self._cancel_requested.add(task_index)
            if self._current_proc and self._current_proc.poll() is None:
                try:
                    self._current_proc.kill()
                except Exception:
                    pass
        task['status'] = 'cancelled'
        self.pipeline_widgets[task_index]['frame'].setVisible(False)
        self._log(f"Dihapus: {task['yt_title']}")


    # ═══════════════════════════════════════════════════════════════════════════
    # PIPELINE ENGINE — Parallel Encode + Upload
    # ═══════════════════════════════════════════════════════════════════════════

    def _start_pipeline(self):
        """Start the pipeline: 1 encode slot + 1 upload slot running in parallel."""
        pending = [i for i, t in enumerate(self.pipeline_queue) if t['status'] == 'waiting']
        if not pending:
            QMessageBox.warning(self, "Kosong", "Tidak ada tugas di antrian!")
            return
        self.btn_start_pipeline.setEnabled(False)
        self.btn_start_pipeline.setText("\u23f3 Pipeline Berjalan...")
        self._log("=" * 30)
        self._log("PIPELINE DIMULAI")
        self._log("=" * 30)
        self._advance_pipeline()

    def _advance_pipeline(self):
        """Check slots and start encode/upload as needed. Called after each task completes."""
        # Start encode if slot is free
        if self._encode_slot is None:
            next_encode = self._find_next('waiting')
            if next_encode is not None:
                self._start_encode(next_encode)

        # Start upload if slot is free
        if self._upload_slot is None:
            next_upload = self._find_next('encoded')
            if next_upload is not None:
                self._start_upload(next_upload)

        # Check if all done
        all_done = all(t['status'] in ('done', 'failed', 'cancelled') for t in self.pipeline_queue)
        if all_done and self._encode_slot is None and self._upload_slot is None:
            self.btn_start_pipeline.setEnabled(True)
            self.btn_start_pipeline.setText("\u25b6 Mulai Pipeline")
            self._log("=" * 30)
            self._log("PIPELINE SELESAI!")
            self._log("=" * 30)
            self._update_slot_indicators()

    def _find_next(self, status):
        for i, t in enumerate(self.pipeline_queue):
            if t['status'] == status and i not in self._cancel_requested:
                return i
        return None

    def _start_encode(self, task_index):
        self._encode_slot = task_index
        task = self.pipeline_queue[task_index]
        task['status'] = 'encoding'
        self._update_task_ui(task_index)
        self._update_slot_indicators()
        self._log(f"[Encode] Mulai: {task['yt_title']}")

        def worker():
            success = self._encode(task, task_index)
            self.signals.encode_done.emit(task_index, success)

        threading.Thread(target=worker, daemon=True).start()

    def _start_upload(self, task_index):
        self._upload_slot = task_index
        task = self.pipeline_queue[task_index]
        task['status'] = 'uploading'
        self._update_task_ui(task_index)
        self._update_slot_indicators()
        self._log(f"[Upload] Mulai: {task['yt_title']}")

        def worker():
            channel_name = task['channel']
            if channel_name not in self.channels:
                youtube, err = YouTubeAuth.get_service(channel_name)
                if not youtube:
                    self.signals.log.emit(f"[Upload] Auth error: {err}")
                    self.signals.upload_done.emit(task_index, False)
                    return
                self.channels[channel_name] = youtube
            youtube = self.channels[channel_name]
            video_id = upload_video_to_youtube(youtube, task, self.signals, task_index)
            self.signals.upload_done.emit(task_index, video_id is not None)

        threading.Thread(target=worker, daemon=True).start()

    def _update_slot_indicators(self):
        if self._encode_slot is not None:
            name = self.pipeline_queue[self._encode_slot]['yt_title'][:20]
            self.lbl_encode_slot.setText(f"\u25CF Encode: {name}...")
            self.lbl_encode_slot.setStyleSheet("font-size: 10px; color: #1D4ED8; font-weight: 600;")
        else:
            self.lbl_encode_slot.setText("\u25CF Encode: idle")
            self.lbl_encode_slot.setStyleSheet("font-size: 10px; color: #6B7280;")

        if self._upload_slot is not None:
            name = self.pipeline_queue[self._upload_slot]['yt_title'][:20]
            self.lbl_upload_slot.setText(f"\u25CF Upload: {name}...")
            self.lbl_upload_slot.setStyleSheet("font-size: 10px; color: #92400E; font-weight: 600;")
        else:
            self.lbl_upload_slot.setText("\u25CF Upload: idle")
            self.lbl_upload_slot.setStyleSheet("font-size: 10px; color: #6B7280;")

    def _update_task_ui(self, task_index):
        if task_index >= len(self.pipeline_widgets):
            return
        w = self.pipeline_widgets[task_index]
        task = self.pipeline_queue[task_index]
        status = task['status']

        if status == 'encoding':
            w['lbl_encode_status'].setText("\U0001F504 Encoding")
            w['lbl_encode_status'].setStyleSheet("font-size: 9px; padding: 2px 5px; border-radius: 8px; background: #DBEAFE; color: #1D4ED8; font-weight: 600;")
            w['frame'].setStyleSheet("QFrame { background-color: #F0F4FF; border: 1px solid #BFDBFE; border-radius: 8px; }")
            w['lbl_num'].setStyleSheet("background-color: #DBEAFE; color: #1D4ED8; border-radius: 11px; font-size: 10px; font-weight: bold;")
        elif status == 'encoded':
            w['lbl_encode_status'].setText("\u2714 Encoded")
            w['lbl_encode_status'].setStyleSheet("font-size: 9px; padding: 2px 5px; border-radius: 8px; background: #D1FAE5; color: #059669; font-weight: 600;")
            w['frame'].setStyleSheet("QFrame { background-color: #F9FAFB; border: 1px solid #E5E7EB; border-radius: 8px; }")
        elif status == 'uploading':
            w['lbl_upload_status'].setText("\u2B06 Uploading")
            w['lbl_upload_status'].setStyleSheet("font-size: 9px; padding: 2px 5px; border-radius: 8px; background: #FEF3C7; color: #92400E; font-weight: 600;")
            w['frame'].setStyleSheet("QFrame { background-color: #FFFBEB; border: 1px solid #FDE68A; border-radius: 8px; }")
            w['lbl_num'].setStyleSheet("background-color: #FEF3C7; color: #92400E; border-radius: 11px; font-size: 10px; font-weight: bold;")
        elif status == 'done':
            w['lbl_upload_status'].setText("\u2714 Selesai")
            w['lbl_upload_status'].setStyleSheet("font-size: 9px; padding: 2px 5px; border-radius: 8px; background: #D1FAE5; color: #059669; font-weight: 600;")
            w['frame'].setStyleSheet("QFrame { background-color: #F0FDF4; border: 1px solid #BBF7D0; border-radius: 8px; }")
            w['lbl_num'].setStyleSheet("background-color: #D1FAE5; color: #059669; border-radius: 11px; font-size: 10px; font-weight: bold;")
        elif status == 'failed':
            w['lbl_encode_status'].setText("\u2716 Gagal")
            w['lbl_encode_status'].setStyleSheet("font-size: 9px; padding: 2px 5px; border-radius: 8px; background: #FEE2E2; color: #DC2626; font-weight: 600;")
            w['frame'].setStyleSheet("QFrame { background-color: #FEF2F2; border: 1px solid #FECACA; border-radius: 8px; }")
            w['lbl_num'].setStyleSheet("background-color: #FEE2E2; color: #DC2626; border-radius: 11px; font-size: 10px; font-weight: bold;")

    # ─── Signal Handlers ─────────────────────────────────────────────────────
    def _on_encode_progress(self, task_index, pct):
        # Could update a progress bar in the widget — for now just pass
        pass

    def _on_upload_progress(self, task_index, pct):
        pass

    def _on_encode_done(self, task_index, success):
        self._encode_slot = None
        task = self.pipeline_queue[task_index]
        if success:
            task['status'] = 'encoded'
            self._log(f"[Encode] Selesai: {task['yt_title']}")
        else:
            task['status'] = 'failed'
            self._log(f"[Encode] GAGAL: {task['yt_title']}")
        self._update_task_ui(task_index)
        self._update_slot_indicators()
        self._advance_pipeline()

    def _on_upload_done(self, task_index, success):
        self._upload_slot = None
        task = self.pipeline_queue[task_index]
        if success:
            task['status'] = 'done'
            self._log(f"[Upload] Selesai: {task['yt_title']}")
        else:
            task['status'] = 'failed'
            self._log(f"[Upload] GAGAL: {task['yt_title']}")
        self._update_task_ui(task_index)
        self._update_slot_indicators()
        self._advance_pipeline()

    def _on_status_changed(self, task_index, new_status):
        if task_index < len(self.pipeline_queue):
            self.pipeline_queue[task_index]['status'] = new_status
            self._update_task_ui(task_index)


    # ═══════════════════════════════════════════════════════════════════════════
    # ENCODING ENGINE (from GoEncoder)
    # ═══════════════════════════════════════════════════════════════════════════

    def _encode(self, tugas, task_index):
        """Full FFmpeg encode pipeline. Runs in worker thread."""
        file_list_audio = None
        crossfade_audio_file = None
        jalur_simpan = tugas['output_path']
        try:
            ffmpeg_exe = imageio_ffmpeg.get_ffmpeg_exe()
            mode_durasi = tugas.get('mode_durasi', 'audio')
            jumlah_loop = tugas.get('jumlah_loop', 1) if mode_durasi == "loop" else 1
            crossfade_enabled = tugas.get('crossfade', False)
            crossfade_dur = tugas.get('crossfade_dur', 2)

            durasi_per_lagu = []
            for jalur_audio in tugas['audio']:
                klip = AudioFileClip(jalur_audio)
                durasi_per_lagu.append(klip.duration)
                klip.close()

            # Build full audio list (with loops)
            full_audio_list = []
            full_durations = []
            for loop_ke in range(jumlah_loop):
                for idx_a, jalur_audio in enumerate(tugas['audio']):
                    full_audio_list.append(jalur_audio)
                    full_durations.append(durasi_per_lagu[idx_a])

            if crossfade_enabled and len(full_audio_list) > 1:
                crossfade_audio_file = tempfile.NamedTemporaryFile(delete=False, suffix='.m4a', dir=os.path.dirname(jalur_simpan) or '.')
                crossfade_audio_file.close()
                crossfade_audio_path = crossfade_audio_file.name
                self.signals.log.emit(f"[Encode] Crossfade audio ({len(full_audio_list)} lagu, {crossfade_dur}s)...")
                sukses_cf = self._build_crossfade_audio(ffmpeg_exe, full_audio_list, full_durations, crossfade_dur, crossfade_audio_path, task_index, tugas.get('audio_bitrate', '256k'))
                if not sukses_cf:
                    self.signals.log.emit("[Encode] Crossfade gagal, fallback tanpa crossfade...")
                    crossfade_enabled = False
                else:
                    total_durasi = sum(full_durations) - crossfade_dur * (len(full_audio_list) - 1)
                    total_durasi = max(0, total_durasi)
                    if mode_durasi == "manual":
                        total_durasi = tugas.get('durasi_manual', total_durasi)

            if not crossfade_enabled or len(full_audio_list) <= 1:
                file_list_audio = tempfile.NamedTemporaryFile(mode='w+', delete=False, suffix='.txt', encoding='utf-8')
                for jalur_audio in full_audio_list:
                    file_list_audio.write(f"file '{jalur_audio.replace(chr(92), '/')}'\n")
                file_list_audio.close()
                total_durasi = sum(full_durations)
                if mode_durasi == "manual":
                    total_durasi = tugas.get('durasi_manual', total_durasi)

            encoder = tugas.get('encoder', 'libx264')
            self.signals.log.emit(f"[Encode] Encoder: {encoder} | Durasi: {self._format_waktu(total_durasi)}")

            if crossfade_enabled and len(full_audio_list) > 1:
                cmd = self._build_cmd_with_audio_file(tugas, ffmpeg_exe, crossfade_audio_path, total_durasi, jalur_simpan)
            else:
                cmd = self._build_cmd(tugas, ffmpeg_exe, file_list_audio.name, total_durasi, jalur_simpan)

            sukses = self._run_ffmpeg(cmd, total_durasi, task_index)

            # GPU fallback to CPU
            if not sukses and encoder != "libx264":
                self.signals.log.emit("[Encode] GPU gagal, fallback CPU...")
                tugas_cpu = dict(tugas)
                tugas_cpu['encoder'] = 'libx264'
                if crossfade_enabled and len(full_audio_list) > 1:
                    cmd2 = self._build_cmd_with_audio_file(tugas_cpu, ffmpeg_exe, crossfade_audio_path, total_durasi, jalur_simpan)
                else:
                    cmd2 = self._build_cmd(tugas_cpu, ffmpeg_exe, file_list_audio.name, total_durasi, jalur_simpan)
                sukses = self._run_ffmpeg(cmd2, total_durasi, task_index)

            return sukses
        except Exception as e:
            self.signals.log.emit(f"[Encode] ERROR: {e}")
            return False
        finally:
            if file_list_audio and os.path.exists(file_list_audio.name):
                try:
                    os.remove(file_list_audio.name)
                except Exception:
                    pass
            if crossfade_audio_file and os.path.exists(crossfade_audio_file.name):
                try:
                    os.remove(crossfade_audio_file.name)
                except Exception:
                    pass

    def _build_crossfade_audio(self, ffmpeg_exe, audio_list, durations, crossfade_dur, output_path, task_index, audio_bitrate):
        n = len(audio_list)
        if n < 2:
            return False
        cmd = [ffmpeg_exe, "-y"]
        for path in audio_list:
            cmd += ["-i", path]
        filters = []
        if n == 2:
            filters.append(f"[0:a][1:a]acrossfade=d={crossfade_dur}:c1=tri:c2=tri[outa]")
        else:
            filters.append(f"[0:a][1:a]acrossfade=d={crossfade_dur}:c1=tri:c2=tri[a1]")
            for i in range(2, n):
                prev_label = f"[a{i-1}]"
                out_label = "[outa]" if i == n - 1 else f"[a{i}]"
                filters.append(f"{prev_label}[{i}:a]acrossfade=d={crossfade_dur}:c1=tri:c2=tri{out_label}")
        filter_complex = ";".join(filters)
        cmd += ["-filter_complex", filter_complex, "-map", "[outa]", "-c:a", "aac", "-b:a", audio_bitrate, output_path]
        try:
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, universal_newlines=True, encoding='utf-8', errors='ignore', creationflags=subprocess.CREATE_NO_WINDOW)
            pola_time = re.compile(r"time=(\d{2}):(\d{2}):(\d{2}\.\d{2})")
            total_cf_durasi = sum(durations) - crossfade_dur * (n - 1)
            for line in proc.stdout:
                if task_index in self._cancel_requested:
                    proc.kill()
                    return False
                m = pola_time.search(line)
                if m and total_cf_durasi > 0:
                    h, mi, s = m.groups()
                    elapsed = int(h)*3600 + int(mi)*60 + float(s)
                    pct = min(100, int((elapsed / total_cf_durasi) * 50))  # crossfade = first 50%
                    self.signals.encode_progress.emit(task_index, pct)
            proc.wait()
            return proc.returncode == 0
        except Exception:
            return False

    def _build_cmd_with_audio_file(self, tugas, ffmpeg_exe, audio_file_path, total_durasi, jalur_simpan):
        fps = tugas['fps']
        cmd = [ffmpeg_exe, "-y", "-stream_loop", "-1", "-i", tugas['video'], "-i", audio_file_path, "-map", "0:v", "-map", "1:a"]
        vf_filter = self._build_overlay_filter(tugas) if tugas.get('overlay') else None
        if vf_filter:
            cmd += ["-vf", vf_filter]
        self._append_video_encoder_options(cmd, tugas)
        cmd += ["-r", fps, "-c:a", "copy", "-t", str(total_durasi), jalur_simpan]
        return cmd

    def _build_cmd(self, tugas, ffmpeg_exe, file_list_path, total_durasi, jalur_simpan):
        fps = tugas['fps']
        mode_durasi = tugas.get('mode_durasi', 'audio')
        cmd = [ffmpeg_exe, "-y", "-stream_loop", "-1", "-i", tugas['video']]
        if mode_durasi == "manual":
            cmd += ["-stream_loop", "-1"]
        cmd += ["-f", "concat", "-safe", "0", "-i", file_list_path, "-map", "0:v", "-map", "1:a"]
        vf_filter = self._build_overlay_filter(tugas) if tugas.get('overlay') else None
        if vf_filter:
            cmd += ["-vf", vf_filter]
        self._append_video_encoder_options(cmd, tugas)
        audio_bitrate = tugas.get('audio_bitrate', '256k')
        cmd += ["-r", fps, "-c:a", "aac", "-b:a", audio_bitrate, "-t", str(total_durasi), jalur_simpan]
        return cmd

    def _append_video_encoder_options(self, cmd, tugas):
        encoder = tugas.get('encoder', 'libx264')
        mode = tugas.get('kompresi', 'Normal')
        preset = self._compression_preset(mode, encoder, tugas.get('fps'))
        bitrate = f"{preset['bitrate']}k"
        maxrate = f"{preset['maxrate']}k"
        bufsize = f"{preset['bufsize']}k"
        if encoder == "libx264":
            cmd += ["-c:v", "libx264", "-preset", preset['x264_preset'], "-b:v", bitrate, "-maxrate", maxrate, "-bufsize", bufsize]
        elif encoder == "h264_nvenc":
            cmd += ["-c:v", "h264_nvenc", "-preset", preset['nvenc_preset'], "-rc", "vbr", "-cq", preset['cq'], "-b:v", bitrate, "-maxrate", maxrate, "-bufsize", bufsize]
        elif encoder == "h264_amf":
            cmd += ["-c:v", "h264_amf", "-quality", "speed", "-rc", "vbr_peak", "-b:v", bitrate, "-maxrate", maxrate, "-bufsize", bufsize]
        elif encoder == "h264_qsv":
            cmd += ["-c:v", "h264_qsv", "-preset", "veryfast", "-b:v", bitrate, "-maxrate", maxrate, "-bufsize", bufsize]

    def _build_overlay_filter(self, tugas):
        """Generate FFmpeg drawtext overlay filter for Now Playing (2-row centered)."""
        font_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Inter-SemiBold.ttf")
        if hasattr(sys, '_MEIPASS'):
            font_path = os.path.join(sys._MEIPASS, "Inter-SemiBold.ttf")
        if not os.path.exists(font_path):
            return None

        font_ffmpeg = font_path.replace("\\", "/")
        if len(font_ffmpeg) > 1 and font_ffmpeg[1] == ':':
            font_ffmpeg = font_ffmpeg[0] + "\\:" + font_ffmpeg[2:]

        pos = tugas.get('overlay_pos', 'Bottom-Center')
        audio_list = tugas['audio']
        durations = tugas.get('audio_durations', [])
        mode_durasi = tugas.get('mode_durasi', 'audio')
        jumlah_loop = tugas.get('jumlah_loop', 1) if mode_durasi == "loop" else 1
        total_tracks = len(audio_list) * jumlah_loop
        crossfade_dur = tugas.get('crossfade_dur', 0) if tugas.get('crossfade', False) else 0

        box_h = 58
        pad = 24
        row1_y_offset = 12
        row2_y_offset = 36

        # Estimate box width
        max_text_len = 0
        for loop_i in range(jumlah_loop):
            for i, audio_path in enumerate(audio_list):
                nama = os.path.splitext(os.path.basename(audio_path))[0]
                nama = self._strip_track_number(nama)
                nama_clean = re.sub(r"['\";,\\%\[\]]", "", nama)
                text_len = len(f"{total_tracks}/{total_tracks}  {nama_clean}")
                if text_len > max_text_len:
                    max_text_len = text_len
        box_w = max(320, min(600, max_text_len * 8 + 40))

        if "Bottom" in pos:
            box_y_expr = f"ih-{box_h + pad}"
            text_box_y = f"main_h-{box_h + pad}"
        else:
            box_y_expr = str(pad)
            text_box_y = str(pad)

        if "Left" in pos:
            box_x_expr = str(pad)
            text_box_x = str(pad)
        elif "Right" in pos:
            box_x_expr = f"iw-{box_w + pad}"
            text_box_x = f"main_w-{box_w + pad}"
        else:
            box_x_expr = f"(iw-{box_w})/2"
            text_box_x = f"(main_w-{box_w})/2"

        filters = []
        cumulative_time = 0.0
        track_num = 0

        for loop_i in range(jumlah_loop):
            for i, audio_path in enumerate(audio_list):
                dur = durations[i] if i < len(durations) else 0
                if dur <= 0:
                    continue
                track_num += 1
                start_t = cumulative_time
                effective_dur = dur - crossfade_dur if (track_num < total_tracks and crossfade_dur > 0) else dur
                effective_dur = max(0.5, effective_dur)
                end_t = cumulative_time + effective_dur

                nama = os.path.splitext(os.path.basename(audio_path))[0]
                nama = self._strip_track_number(nama)
                nama_clean = re.sub(r"['\";,:\\%\[\]]", "", nama)
                display_text = f"{track_num}/{total_tracks}  {nama_clean}"
                enable = f"between(t,{start_t:.2f},{end_t:.2f})"

                filters.append(f"drawbox=x={box_x_expr}:y={box_y_expr}:w={box_w}:h={box_h}:color=black@0.6:t=fill:enable='{enable}'")
                text_x = f"{text_box_x}+({box_w}-text_w)/2"
                text_y = f"{text_box_y}+{row1_y_offset}"
                filters.append(f"drawtext=fontfile='{font_ffmpeg}':text='{display_text}':fontsize=15:fontcolor=white:x={text_x}:y={text_y}:enable='{enable}'")

                dur_min = int(effective_dur // 60)
                dur_sec = int(effective_dur % 60)
                dur_str = f"{dur_min:02d}\\:{dur_sec:02d}"
                elapsed_x = f"{text_box_x}+({box_w}-text_w)/2"
                elapsed_y = f"{text_box_y}+{row2_y_offset}"
                elapsed_expr = f"%{{eif\\:floor((t-{start_t:.2f})/60)\\:d\\:2}}\\:%{{eif\\:mod(t-{start_t:.2f}\\,60)\\:d\\:2}}"
                filters.append(f"drawtext=fontfile='{font_ffmpeg}':text='{elapsed_expr} / {dur_str}':fontsize=12:fontcolor=white@0.6:x={elapsed_x}:y={elapsed_y}:enable='{enable}'")

                cumulative_time = end_t

        return ",".join(filters) if filters else None

    def _run_ffmpeg(self, cmd, total_durasi, task_index):
        try:
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, universal_newlines=True, encoding='utf-8', errors='ignore', creationflags=subprocess.CREATE_NO_WINDOW)
            self._current_proc = proc
            pola = re.compile(r"time=(\d{2}):(\d{2}):(\d{2}\.\d{2})")
            for line in proc.stdout:
                if task_index in self._cancel_requested:
                    proc.kill()
                    return False
                m = pola.search(line)
                if m:
                    h, mi, s = m.groups()
                    elapsed = int(h)*3600 + int(mi)*60 + float(s)
                    if total_durasi > 0:
                        pct = min(100.0, (elapsed / total_durasi) * 100)
                        self.signals.encode_progress.emit(task_index, pct)
            proc.wait()
            self._current_proc = None
            return proc.returncode == 0
        except Exception:
            self._current_proc = None
            return False


    # ═══════════════════════════════════════════════════════════════════════════
    # LOGGING & UTILITIES
    # ═══════════════════════════════════════════════════════════════════════════

    LOG_COLORS = {
        "info": "#94A3B8", "auth": "#818CF8", "queue": "#38BDF8",
        "progress": "#5EEAD4", "success": "#34D399", "warning": "#FBBF24",
        "error": "#F87171", "separator": "#475569", "encode": "#93C5FD",
        "upload": "#FBBF24",
    }

    def _log(self, msg):
        category = self._detect_log_category(msg)
        waktu = datetime.datetime.now().strftime("%H:%M:%S")
        color = self.LOG_COLORS.get(category, "#94A3B8")
        self.log_area.append(f"<span style='color:#6B7280;'>[{waktu}]</span> <span style='color:{color};'>{msg}</span>")
        if hasattr(self, '_log_file') and self._log_file:
            try:
                self._log_file.write(f"[{waktu}] {msg}\n")
                self._log_file.flush()
            except Exception:
                pass

    def _on_log(self, msg):
        self._log(msg)

    def _detect_log_category(self, msg):
        msg_lower = msg.lower()
        if "error" in msg_lower or "gagal" in msg_lower:
            return "error"
        elif "selesai" in msg_lower or "berhasil" in msg_lower:
            return "success"
        elif "[encode]" in msg_lower:
            return "encode"
        elif "[upload]" in msg_lower:
            return "upload"
        elif "pipeline" in msg_lower or "antrian" in msg_lower:
            return "queue"
        elif msg.startswith("="):
            return "separator"
        elif "auth" in msg_lower or "login" in msg_lower:
            return "auth"
        else:
            return "info"

    @staticmethod
    def _format_waktu(detik):
        if detik is None:
            return "00:00:00"
        jam, sisa = divmod(int(detik), 3600)
        menit, det = divmod(sisa, 60)
        return f"{jam:02d}:{menit:02d}:{det:02d}"

    # ─── Reset ───────────────────────────────────────────────────────────────
    def _reset_all(self):
        # Check if any task is actively processing
        if self._encode_slot is not None or self._upload_slot is not None:
            QMessageBox.warning(self, "Peringatan", "Tidak bisa reset saat pipeline berjalan!")
            return

        self.jalur_video = ""
        self.lbl_video.setText("Belum dipilih")
        self.lbl_video.setStyleSheet("font-size: 11px; color: #9CA3AF;")
        self.video_preview.setPixmap(QPixmap())
        self.video_preview.setText("Drop atau pilih video")
        self.video_preview.setStyleSheet("QLabel { background-color: #F7FAFA; color: #9CA3AF; border: 2px dashed #D1D5DB; border-radius: 10px; font-size: 11px; }")

        self.daftar_audio.clear()
        self.audio_durations.clear()
        self._refresh_audio_list()

        # Clear pipeline queue
        self.pipeline_queue.clear()
        self._cancel_requested.clear()
        for w in self.pipeline_widgets:
            w['frame'].deleteLater()
        self.pipeline_widgets.clear()

        self.log_area.clear()
        self.tracklist_area.clear()

        # Reset duration
        self.radio_audio.setChecked(True)
        self.entry_jam.setText("00")
        self.entry_menit.setText("00")
        self.entry_detik.setText("00")
        self.entry_loop.setText("2")

        # Reset render options
        self.chk_overlay.setChecked(False)
        self.combo_overlay_pos.setCurrentIndex(1)
        self.combo_bitrate.setCurrentIndex(1)
        self.chk_crossfade.setChecked(False)
        self.combo_crossfade_dur.setCurrentIndex(1)

        # Reset YouTube fields
        self.entry_title.clear()
        self.entry_desc.clear()
        self.entry_tags.clear()
        self.combo_privacy.setCurrentIndex(2)
        self.chk_schedule.setChecked(False)
        self.combo_playlist.setCurrentIndex(0)
        self._thumbnail_path = None
        self.lbl_thumb.setText("(opsional)")
        self.lbl_thumb.setStyleSheet("font-size: 10px; color: #9CA3AF;")

        # Reset Excel
        self._excel_data = None
        self._excel_seo_data = {}
        self._excel_seo_ordered = []
        self._excel_folder = ""
        self._excel_matches = {}
        self._set_audio_mode('manual')
        self.excel_import_zone.setStyleSheet(
            "QPushButton { background: #F0FDFA; border: 2px dashed #0D9488; border-radius: 10px; padding: 14px; font-size: 11px; font-weight: 600; color: #6B7280; }"
            "QPushButton:hover { border-color: #0B7C72; background: #CCFBF1; }"
        )
        self.excel_import_zone.setText("\U0001F4CA  Klik untuk Import Excel (.xlsx)")
        self.combo_variation.clear()
        self.lbl_variation_note.setText("")
        self.lbl_excel_folder.setText("Belum dipilih")
        self.lbl_match_status.setText("")
        self.excel_variation_widget.setVisible(False)

        # Reset slot indicators
        self._encode_slot = None
        self._upload_slot = None
        self._update_slot_indicators()
        self.btn_start_pipeline.setEnabled(True)
        self.btn_start_pipeline.setText("\u25b6 Mulai Pipeline")

        self._update_size_estimate()
        self._log("Reset — semua input dibersihkan.")


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    force_light_palette(app)
    app.setStyleSheet(STYLESHEET)
    window = GoStudioWindow()
    window.show()
    sys.exit(app.exec())

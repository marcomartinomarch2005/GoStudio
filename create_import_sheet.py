"""
Generate GoStudio_Import.xlsx — a single Excel file optimized for GoStudio import.
Combines Sequence Variations + Visual SEO into one clean structure.

Sheet 1: "Sequence Variations" — track order per variation (for audio ordering)
Sheet 2: "Visual SEO" — YouTube metadata per variation (for auto-fill)

Both sheets use "Variation Name" as the shared key.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Sequence Variations
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sequence Variations"

# Header row — matches GoStudio _COL_ALIASES exactly
headers_seq = ["Variation Name", "Track Order", "Title", "Session", "Mood", "Strategic Note"]
ws1.append(headers_seq)

# Style header
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
for col in range(1, len(headers_seq) + 1):
    cell = ws1.cell(row=1, column=col)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Data — 10 variations x 20 tracks each
variations_data = {
    "Canonical Mid-Session Lift": {
        "mood": "Deep work / baseline positioning",
        "tracks": [
            "Signal Bloom", "Intro to Pulse", "Quiet Circuit", "Neon Delay",
            "Workstream Gate", "Focus Wire", "Deep Current", "Code Horizon",
            "Glass Machine", "Odd Pulse Soft", "Cinematic Rise", "Focus Engine Peak",
            "After Peak Current", "Night Work Mode", "Quiet Systems", "Return to Flow",
            "Night Workspace", "Final Riff Glow", "Soft Resolution", "Loopable Aftertone"
        ]
    },
    "Coding Momentum Flow": {
        "mood": "Coding / software engineering",
        "tracks": [
            "Signal Bloom", "Quiet Circuit", "Neon Delay", "Intro to Pulse",
            "Workstream Gate", "Focus Wire", "Code Horizon", "Odd Pulse Soft",
            "Deep Current", "Glass Machine", "Cinematic Rise", "Focus Engine Peak",
            "After Peak Current", "Night Work Mode", "Return to Flow", "Quiet Systems",
            "Night Workspace", "Final Riff Glow", "Soft Resolution", "Loopable Aftertone"
        ]
    },
    "Executive Deep Focus": {
        "mood": "Executive focus / strategy work",
        "tracks": [
            "Signal Bloom", "Intro to Pulse", "Neon Delay", "Quiet Circuit",
            "Deep Current", "Glass Machine", "Quiet Systems", "Code Horizon",
            "After Peak Current", "Cinematic Rise", "Focus Engine Peak", "Night Work Mode",
            "Return to Flow", "Odd Pulse Soft", "Focus Wire", "Workstream Gate",
            "Night Workspace", "Final Riff Glow", "Soft Resolution", "Loopable Aftertone"
        ]
    },
    "Precision Prog Lab": {
        "mood": "Prog enthusiast / technical focus",
        "tracks": [
            "Signal Bloom", "Intro to Pulse", "Focus Wire", "Quiet Circuit",
            "Workstream Gate", "Neon Delay", "Code Horizon", "Odd Pulse Soft",
            "Deep Current", "Glass Machine", "Focus Engine Peak", "Cinematic Rise",
            "After Peak Current", "Night Work Mode", "Quiet Systems", "Return to Flow",
            "Night Workspace", "Soft Resolution", "Final Riff Glow", "Loopable Aftertone"
        ]
    },
    "Hypebeast Night Sanctuary": {
        "mood": "Cinematic focus / luxury man cave",
        "tracks": [
            "Signal Bloom", "Intro to Pulse", "Neon Delay", "Workstream Gate",
            "Quiet Circuit", "Focus Wire", "Deep Current", "Glass Machine",
            "Cinematic Rise", "Focus Engine Peak", "After Peak Current", "Night Work Mode",
            "Code Horizon", "Odd Pulse Soft", "Quiet Systems", "Return to Flow",
            "Night Workspace", "Final Riff Glow", "Soft Resolution", "Loopable Aftertone"
        ]
    },
    "Study Reading Clean Flow": {
        "mood": "Studying / reading / calm work",
        "tracks": [
            "Signal Bloom", "Intro to Pulse", "Neon Delay", "Quiet Circuit",
            "Glass Machine", "Deep Current", "Quiet Systems", "Code Horizon",
            "Cinematic Rise", "After Peak Current", "Night Work Mode", "Focus Engine Peak",
            "Return to Flow", "Odd Pulse Soft", "Focus Wire", "Workstream Gate",
            "Night Workspace", "Soft Resolution", "Final Riff Glow", "Loopable Aftertone"
        ]
    },
    "Cinematic Reading Build": {
        "mood": "Cinematic reading / creative flow",
        "tracks": [
            "Signal Bloom", "Intro to Pulse", "Neon Delay", "Quiet Circuit",
            "Deep Current", "Glass Machine", "Night Work Mode", "Quiet Systems",
            "Cinematic Rise", "After Peak Current", "Code Horizon", "Focus Engine Peak",
            "Odd Pulse Soft", "Return to Flow", "Focus Wire", "Workstream Gate",
            "Night Workspace", "Final Riff Glow", "Soft Resolution", "Loopable Aftertone"
        ]
    },
    "Odd-Time Focus Arc": {
        "mood": "Odd-time focus / prog niche",
        "tracks": [
            "Signal Bloom", "Intro to Pulse", "Quiet Circuit", "Focus Wire",
            "Neon Delay", "Workstream Gate", "Odd Pulse Soft", "Code Horizon",
            "Deep Current", "Glass Machine", "Cinematic Rise", "Focus Engine Peak",
            "After Peak Current", "Night Work Mode", "Return to Flow", "Quiet Systems",
            "Night Workspace", "Soft Resolution", "Final Riff Glow", "Loopable Aftertone"
        ]
    },
    "Late Night Work Mode": {
        "mood": "Night work / remote work",
        "tracks": [
            "Signal Bloom", "Neon Delay", "Intro to Pulse", "Quiet Circuit",
            "Workstream Gate", "Deep Current", "Code Horizon", "Glass Machine",
            "Cinematic Rise", "Focus Engine Peak", "After Peak Current", "Night Work Mode",
            "Quiet Systems", "Odd Pulse Soft", "Return to Flow", "Focus Wire",
            "Night Workspace", "Final Riff Glow", "Soft Resolution", "Loopable Aftertone"
        ]
    },
    "Calm Recovery Loop": {
        "mood": "Recovery focus / repeat listening",
        "tracks": [
            "Signal Bloom", "Intro to Pulse", "Neon Delay", "Quiet Circuit",
            "Deep Current", "Quiet Systems", "Glass Machine", "Code Horizon",
            "Cinematic Rise", "After Peak Current", "Night Work Mode", "Return to Flow",
            "Focus Engine Peak", "Odd Pulse Soft", "Focus Wire", "Workstream Gate",
            "Night Workspace", "Soft Resolution", "Final Riff Glow", "Loopable Aftertone"
        ]
    },
}

for var_name, var_info in variations_data.items():
    for order, track_title in enumerate(var_info["tracks"], 1):
        ws1.append([var_name, order, track_title, "", var_info["mood"], ""])

# Autofit columns
for col in ws1.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws1.column_dimensions[col_letter].width = min(max_length + 2, 40)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Visual SEO
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Visual SEO")

# Header row — matches GoStudio _SEO_COL_ALIASES exactly
headers_seo = ["Variation Name", "Title", "Description + Hashtags", "Tags"]
ws2.append(headers_seo)

for col in range(1, len(headers_seo) + 1):
    cell = ws2.cell(row=1, column=col)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# SEO data per variation — same order as Sequence Variations
seo_data = [
    {
        "variation": "Canonical Mid-Session Lift",
        "title": "Progressive Rock Music for Deep Work | 1 Hour Instrumental Focus Session",
        "description": "Use this progressive rock music for deep work, coding, studying, reading, and creative focus. This 1 hour instrumental focus session is built with evolving riffs, atmospheric guitars, subtle odd-time grooves, cinematic textures, and no vocals, designed to keep your mind engaged without pulling attention away from the task.\n\nVelvet Riff creates instrumental progressive rock focus music for long concentration sessions: no lyrics, no distractions, just guitar-driven flow state energy for work, programming, study, reading, and creative productivity.\n\n#ProgressiveRock #FocusMusic #InstrumentalRock",
        "tags": "progressive rock music, progressive rock for deep work, instrumental progressive rock, instrumental prog rock, deep work music, focus music, guitar focus music, no vocals music, no lyrics music, prog rock for coding, study music with guitar, work concentration music, flow state music, productivity music, atmospheric guitar music, cinematic prog rock, clean guitar focus, Velvet Riff",
    },
    {
        "variation": "Coding Momentum Flow",
        "title": "Instrumental Prog Rock for Coding | 60 Minute Flow State Music",
        "description": "A 60 minute instrumental prog rock session for coding, programming, debugging, and deep technical work. Built with steady guitar riffs, tight drums, warm bass, atmospheric textures, and no vocals, this mix is designed to support flow state without distracting your thinking.\n\nUse it for software development, problem solving, focused work blocks, study, writing, and creative productivity.\n\n#CodingMusic #ProgRock #FocusMusic",
        "tags": "instrumental prog rock for coding, coding music, prog rock for coding, programming music, developer music, flow state music, focus music, no vocals music, guitar focus music, progressive rock instrumental, work music, productivity music, deep work music, Velvet Riff",
    },
    {
        "variation": "Executive Deep Focus",
        "title": "Progressive Rock for Work and Concentration | 1 Hour Executive Focus",
        "description": "A premium instrumental progressive rock focus session for deep work, strategy, planning, writing, and high-level concentration. Built with evolving guitar riffs, atmospheric textures, controlled energy, and no vocals.\n\nUse this when you need clarity, control, and long-form focus without lyrics or distractions.\n\n#DeepWork #ProgressiveRock #FocusMusic",
        "tags": "progressive rock for work and concentration, executive focus music, deep work music, instrumental prog rock, no vocals music, strategy music, work music, productivity music, focus riffs, Velvet Riff",
    },
    {
        "variation": "Precision Prog Lab",
        "title": "Focus Riffs Vol. 1 | Instrumental Progressive Rock for Productivity",
        "description": "Focus Riffs Vol. 1 is an instrumental progressive rock productivity session built for detail work, creative problem solving, and long concentration. Expect evolving riffs, steady grooves, atmospheric guitars, controlled dynamics, and no vocals.\n\nFor deep work, coding, writing, design, study, and focused technical tasks.\n\n#FocusRiffs #InstrumentalRock #Productivity",
        "tags": "focus riffs, instrumental progressive rock for productivity, prog rock focus music, technical focus music, guitar focus music, no vocals music, odd time focus music, work music, Velvet Riff",
    },
    {
        "variation": "Hypebeast Night Sanctuary",
        "title": "Flow State Rock Music | 1 Hour Instrumental Prog Focus",
        "description": "Flow state rock music for long work sessions, creative focus, and private deep concentration. This instrumental progressive rock session blends evolving riffs, atmospheric guitars, steady drums, and cinematic textures with no vocals or lyrics.\n\nTurn the world off and stay in the zone.\n\n#FlowState #InstrumentalRock #FocusMusic",
        "tags": "flow state rock music, instrumental prog focus, progressive rock focus music, no vocals rock music, deep work music, cinematic prog rock, focus riffs, Velvet Riff",
    },
    {
        "variation": "Study Reading Clean Flow",
        "title": "No Lyrics Progressive Rock for Studying | Guitar Focus Music",
        "description": "No lyrics progressive rock for studying, reading, note-taking, and long concentration. This guitar focus music session uses clean progressive riffs, atmospheric textures, warm bass, and controlled drums to keep the mind engaged without vocals or lyrical distraction.\n\nBest used for study blocks, reading sessions, writing, research, and focused learning.\n\n#StudyMusic #ProgressiveRock #NoLyrics",
        "tags": "no lyrics progressive rock, progressive rock for studying, study music with guitar, guitar focus music, instrumental rock study music, no vocals focus music, reading music, concentration music, clean guitar study, progressive rock instrumental, Velvet Riff",
    },
    {
        "variation": "Cinematic Reading Build",
        "title": "Cinematic Prog Rock for Reading | No Vocals Focus Music",
        "description": "A cinematic instrumental progressive rock focus session for reading, writing, creative thinking, and long concentration. Built with atmospheric guitars, clean-to-crunch dynamics, subtle build-ups, and no vocals.\n\nBest for reading, creative planning, editing, design, and quiet productivity.\n\n#CinematicRock #ReadingMusic #FocusMusic",
        "tags": "cinematic prog rock for reading, no vocals focus music, instrumental progressive rock, atmospheric guitar focus music, reading music, creative focus music, guitar music for work, Velvet Riff",
    },
    {
        "variation": "Odd-Time Focus Arc",
        "title": "Odd Time Focus Music | Progressive Rock Study Session",
        "description": "Odd time focus music for listeners who want progressive rhythm without losing concentration. This instrumental prog rock session uses subtle odd-time grooves, evolving guitars, warm bass, and controlled drums designed for study, work, and focused creative flow.\n\n#OddTime #ProgressiveRock #FocusMusic",
        "tags": "odd time focus music, progressive rock study session, instrumental prog rock, odd time groove, technical focus music, no vocals prog rock, focus music, study music with guitar, Velvet Riff",
    },
    {
        "variation": "Late Night Work Mode",
        "title": "Progressive Guitar Music for Concentration | Night Work Mode",
        "description": "Progressive guitar music for night work, concentration, and long creative sessions. This no-vocal instrumental prog rock mix uses steady riffs, atmospheric guitars, warm bass, and controlled dynamics for deep focus after dark.\n\n#NightWork #ProgressiveRock #FocusMusic",
        "tags": "progressive guitar music for concentration, night work music, instrumental prog rock, deep focus music, no vocals music, guitar focus music, productivity music, Velvet Riff",
    },
    {
        "variation": "Calm Recovery Loop",
        "title": "Guitar Focus Music for Work | Instrumental Progressive Rock Session",
        "description": "A guitar-driven instrumental progressive rock session for work, concentration, and creative flow. Built with evolving riffs, clean and crunch guitar textures, warm bass, steady drums, cinematic ambience, and no vocals.\n\nUse it as background music for deep work, writing, coding, design, editing, reading, and productivity.\n\n#GuitarFocus #InstrumentalRock #WorkMusic",
        "tags": "guitar focus music, instrumental progressive rock, guitar music for work, work music, concentration music, progressive rock music, no vocals music, focus riffs, deep work music, creative flow music, Velvet Riff",
    },
]

for entry in seo_data:
    ws2.append([entry["variation"], entry["title"], entry["description"], entry["tags"]])

# Autofit
for col in ws2.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, min(len(str(cell.value)), 60))
    ws2.column_dimensions[col_letter].width = min(max_length + 2, 60)


# ═══════════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════════
output_path = r"C:\Users\Pongo Studio\Documents\Builder\gostudio\GoStudio_Import_Ready.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
print(f"Sheet 1: 'Sequence Variations' — {len(variations_data)} variations x 20 tracks")
print(f"Sheet 2: 'Visual SEO' — {len(seo_data)} SEO packages")
print()
print("KEY DESIGN:")
print("- 'Variation Name' is the SHARED KEY between both sheets")
print("- Sheet 1 has 'Track Order' column (triggers priority 1 detection)")
print("- Sheet 2 has 'Title' + 'Description + Hashtags' + 'Tags'")
print("- No numeric Variation column = no ambiguity")
